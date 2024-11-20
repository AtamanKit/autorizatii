from flask import Flask, flash, jsonify, request, redirect, url_for
from werkzeug.utils import secure_filename
from flask_pymongo import PyMongo
from flask_cors import CORS

from fpdf import FPDF

from openpyxl import load_workbook
from openpyxl.styles import Alignment

import datetime

import psycopg2
from psycopg2.extensions import ISOLATION_LEVEL_AUTOCOMMIT

import pandas as pd
import re
import json
import os
import random
import io
# import direnv
from dotenv import load_dotenv


#Working with google API
import pickle
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from apiclient.http import MediaFileUpload, MediaIoBaseDownload

def date_get():
    myDate = datetime.datetime.now()
    yearNow = myDate.strftime('%Y')
    # yearNow = "2021"
    mthNow = myDate.strftime('%m')
    dayNow = myDate.strftime('%d')
    hourNow = myDate.strftime('%H')
    minuteNow = myDate.strftime('%M')
    yearNowShort = myDate.strftime('%y')
    # yearNowShort = "21"

    return (
        yearNow,
        mthNow,
        dayNow,
        hourNow,
        minuteNow,
        yearNowShort,
        myDate
    )

app = Flask(__name__)

# direnv.load()
# Load environment variables from .env file
load_dotenv()

app.config['MONGO_URI'] = f'mongodb://{os.environ["MONGO_USERNAME"]}:{os.environ["MONGO_PASSWORD"]}@db/djUN_{date_get()[5]}?authSource=admin'
mongo = PyMongo(app)
app.config['MONGO_URI'] = f'mongodb://{os.environ["MONGO_USERNAME"]}:{os.environ["MONGO_PASSWORD"]}@db/General?authSource=admin'
mgGen = PyMongo(app)

# print("Begin_test")

# UPLOAD_FOLDER = 'C:\Red-Nord\Autorizatii'
# ALLOWED_EXTENSIONS = {'txt', 'pdf', 'png', 'jpg', 'jpeg', 'gif'}

SAIDI_LOCATION = "T:/DISPECERAT CENTRAL/DECONECTARI ZILNICE/RETELE 04kV/UN/"
# SAIDI_LOCATION = "D:/Red-Nord/DISC T/DISPECERAT CENTRAL/DECONECTARI ZILNICE/Retele 04kV/UN/"

REGAL_LOCATION = "T:/DISPECERAT CENTRAL/REGISTRU AUTORIZATII PDJT/UN/"
# REGAL_LOCATION = "D:/Red-Nord/DISC T/DISPECERAT CENTRAL/REGISTRU AUTORIZATII PDJT/UN/"

# LOCAL_DIRECTORY = "C:/Users/Public/Biroul Dispecerului/"
UPLOAD_DIRECTORY = os.path.join(os.path.dirname(__file__), "tmp")

OFICII = [
    {"name": "Ungheni", "abr": "UN"},
    {"name": "Falesti", "abr": "FL"},
    {"name": "Glodeni", "abr": "GL"},
    {"name": "Riscani", "abr": "RS"}
]

CORS(app, origins=[
            'https://rednord.org',
            'http://localhost:3000',
            'http://localhost:3001',
            'https://bir-react.firebaseapp.com'
        ]
     )

# try:
#     conn = psycopg2.connect(
#             host = f'{os.environ["POSTGRES_HOST"]}',
#             database = f'ungheni_{date_get()[5]}_test',
#             user = f'{os.environ["POSTGRES_USER"]}',
#             password = f'{os.environ["POSTGRES_PASSWORD"]}'
#         )
#     cur = conn.cursor()
# except psycopg2.OperationalError:
#     conn = psycopg2.connect(
#         host = f'{os.environ["POSTGRES_HOST"]}',
#         user = f'{os.environ["POSTGRES_USER"]}',
#         password = f'{os.environ["POSTGRES_PASSWORD"]}'
#     )
#     conn.set_isolation_level(ISOLATION_LEVEL_AUTOCOMMIT)
#     cur = conn.cursor()
#     cur.execute(f'CREATE DATABASE ungheni_{date_get()[5]}_test')


@app.route('/autorizatii/<abrOf>/<monthNum>/<pt>/<angaj>/')
def get_al_of(abrOf, monthNum, pt, angaj):
    output = []
    myTotal = ""
    reg_al = "al_" + date_get()[0] + "_" + monthNum

    for i in mongo.db[reg_al].find().sort("_id", -1).limit(1):
        myTotal = i["_id"]
    # output.append(myTotal)

    if abrOf != "AD":
        if pt == "undefined" and angaj == "undefined":
            for i in mongo.db[reg_al].find({"oficiul": abrOf}):
                output.append(i)
        elif pt != "undefeined" and angaj == "undefined":
            for i in mongo.db[reg_al].find({"oficiul": abrOf, "pt": pt}):
                output.append(i)
        elif pt != "undefined" and angaj != "undefined":
            for i in mongo.db[reg_al].find({"oficiul": abrOf, "pt": pt, "sef": angaj}):
                output.append(i)
        elif pt == "undefined" and angaj != "undefined":
            for i in mongo.db[reg_al].find({"oficiul": abrOf, "sef": angaj}):
                output.append(i)
    else:
        if pt == "undefined" and angaj == "undefined":
            for i in mongo.db[reg_al].find():
                output.append(i)
        elif pt != "undefeined" and angaj == "undefined":
            for i in mongo.db[reg_al].find({"pt": pt}):
                output.append(i)
        elif pt != "undefined" and angaj != "undefined":
            for i in mongo.db[reg_al].find({"pt": pt, "sef": angaj}):
                output.append(i)
        elif pt == "undefined" and angaj != "undefined":
            for i in mongo.db[reg_al].find({"sef": angaj}):
                output.append(i)

    # print(output)
    if output == []:
        output = [{
            "_id": "",
            "oficiul": "",
            "nr_ds": "",
            "nr_al": "",
            "cu_dec": "",
            "instalatia": "",
            "pt": "",
            "fid_nr": "",
            "localitatea": "",
            "lucrarile": "",
            "sef": "",
            "mem_ech": "",
            "emitent": "",
            "mas_teh": "",
            "semnatura": "",
            "starea": "",
            "pregatire": "",
            "admitere": "",
            "terminare": "",
            "link": "",
        }]

    output = [
        output,
        myTotal
    ]

    return jsonify(output)

@app.route('/pt/<abrOf>/')
def get_pt(abrOf):
    output = []
    myPt = "pt_" + abrOf.lower()
    for i in mgGen.db[myPt].find():
        output.append(i)

    output = sorted(output, key=lambda k: k['pt'])
    return jsonify(output)

@app.route('/instalatia/')
def get_inst():
    output = []
    for i in mgGen.db.nomenclator.find({"name": "instalatia"}):
        output.append(i)

    return jsonify(output)

@app.route('/lucrari/')
def get_lucrari():
    output = []
    for i in mgGen.db.nomenclator.find({"name": "lucrari"}):
        output.append(i)

    return jsonify(output)

@app.route('/cauza-dec/')
def get_cauza_dec():
    output = []
    for i in mgGen.db.nomenclator.find({"name": "cauza_dec"}):
        output.append(i)

    return jsonify(output)

@app.route('/cauza-sig/')
def get_couza_sig():
    output = []
    for i in mgGen.db.nomenclator.find({"name": "cauza_sig"}):
        output.append(i)

    output[0]['lista'].sort()

    return jsonify(output)

@app.route('/cu_dec/')
def get_cu_dec():
    output = []
    for i in mgGen.db.nomenclator.find({"name": "cu_dec"}):
        output.append(i)

    return jsonify(output)

@app.route('/cerinte_ts')
def get_cerinte_ts():
    output = []
    for i in mgGen.db.nomenclator.find({"name": "cerinte_ts"}):
        output.append(i)
    return jsonify(output)

@app.route('/deranjament/vazut/<deranjNum>/', methods=['PUT'])
def get_deranj_vazut(deranjNum):
    reg_deranj = "deranj_" + date_get()[0]
    mongo.db[reg_deranj].update_one({"_id": int(deranjNum)}, {
        "$set": {"responsabil": request.get_json()["vazut"]}
    })
    return "success"

@app.route('/deranjament/executat/<deranjNum>/', methods=['PUT'])
def get_deranj_exec(deranjNum):
    reg_deranj = "deranj_" + date_get()[0]
    mongo.db[reg_deranj].update_one({"_id": int(deranjNum)}, {
        "$set": {"starea": request.get_json()["exec"]}
    })
    return "success"

@app.route('/autorizatii/starea/<dsAlId>/', methods=['PUT'])
def get_accept(dsAlId):
    starea_list = request.get_json()["starea"].split(".")
    reg_al = "al_" + date_get()[0] + "_" + starea_list[1]
    if mongo.db[reg_al].find_one({"_id": int(dsAlId)})["starea"] == "Nou inregistrat":
        mongo.db[reg_al].update_one({"_id": int(dsAlId)}, {
            "$set": {"starea": request.get_json()["starea"]}
        })
        return "approved"
    else:
        return "non-approved"

@app.route('/autorizatii/pregatire/<dsAlId>/', methods=['PUT'])
def get_preg(dsAlId):
    starea_list = request.get_json()["starea"].split(".")
    # print(starea_list)
    reg_al = "al_" + date_get()[0] + "_" + starea_list[1]
    if mongo.db[reg_al].find_one({"_id": int(dsAlId)})["pregatire"] == "Pregatire":
        mongo.db[reg_al].update_one({"_id": int(dsAlId)}, {
            "$set": {"pregatire": request.get_json()["pregatire"]}})
        return "approved"
    else:
        return "non-approved"

@app.route('/autorizatii/admitere/<dsAlId>/', methods=['PUT'])
def get_adm(dsAlId):
    starea_list = request.get_json()["starea"].split(".")
    reg_al = "al_" + date_get()[0] + "_" + starea_list[1]
    if mongo.db[reg_al].find_one({"_id": int(dsAlId)})["admitere"] == "Admitere":
        mongo.db[reg_al].update_one({"_id": int(dsAlId)}, {
            "$set": {"admitere": request.get_json()["admitere"]}})
        return "approved"
    else:
        return "non-approved"

@app.route('/autorizatii/terminare/<dsAlId>/', methods=['PUT'])
def get_term(dsAlId):
    starea_list = request.get_json()["starea"].split(".")
    reg_al = "al_" + date_get()[0] + "_" + starea_list[1]
    if mongo.db[reg_al].find_one({"_id": int(dsAlId)})["terminare"] == "Terminare" or \
        re.search("Intrerupere:", mongo.db[reg_al].find_one({"_id": int(dsAlId)})["terminare"]):
        mongo.db[reg_al].update_one({"_id": int(dsAlId)}, {
            "$set": {
                    "pregatire": request.get_json()["pregatire"],
                    "admitere": request.get_json()["admitere"],
                    "terminare": request.get_json()["terminare"]
            }})

        if mongo.db[reg_al].find_one({"_id": int(dsAlId)})["nr_al"] != "":
            #---------------------Necesary Variables From Al---------------------
            oficiul = mongo.db[reg_al].find_one({"_id": int(dsAlId)})["oficiul"]
            cu_dec = mongo.db[reg_al].find_one({"_id": int(dsAlId)})["cu_dec"]
            pt = mongo.db[reg_al].find_one({"_id": int(dsAlId)})["pt"]
            fid = mongo.db[reg_al].find_one({"_id": int(dsAlId)})["fid_nr"]
            localitatea = mongo.db[reg_al].find_one({"_id": int(dsAlId)})["localitatea"]
            lucrarile = mongo.db[reg_al].find_one({"_id": int(dsAlId)})["lucrarile"]

            myPreg = mongo.db[reg_al].find_one({"_id": int(dsAlId)})["pregatire"]
            myTerm = mongo.db[reg_al].find_one({"_id": int(dsAlId)})["terminare"]

            #----------------Working with time and deltatime----------------------------
            myPreg_list = myPreg.split("\n")
            myPreg_date_str = myPreg_list[1] + " " + myPreg_list[2]
            myPreg_date = datetime.datetime.strptime(myPreg_date_str, "%d.%m.%y %H:%M")

            myTerm_list = myTerm.split("\n")
            myTerm_date_str = myTerm_list[1] + " " + myTerm_list[2]
            myTerm_date = datetime.datetime.strptime(myTerm_date_str, "%d.%m.%y %H:%M")

            delta_preg_term = myTerm_date - myPreg_date
            delta_preg_term = str(delta_preg_term)

            #-----------------Working with "localitatea", nr.cons-----------------------------
            reg_pt = "pt_" + oficiul.lower()

            for i in mgGen.db[reg_pt].find():
                if pt == i["pt"]:
                    totNrCas = i["total_cas"]
                    fidNrCas = round(totNrCas/3)
                    if fidNrCas > 65:
                        fidNrCas = random.randrange(60, 70)

                    myComma = re.search(",", fid)
                    if myComma:
                        myNumber = fid.count(",")
                        fidNrCas = (myNumber + 1) * fidNrCas

                    totNrEc = i["total_non_cas"]
                    if totNrEc <= 2 and totNrCas == 0:
                        fidNrEc = 1
                    elif totNrEc > 12:
                        fidNrEc = random.randrange(2, 5)
                    else:
                        fidNrEc = round(totNrEc/3)

                    myComma = re.search(",", fid)
                    if myComma:
                        myNumber = fid.count(",")
                        fidNrEc = (myNumber + 1) * fidNrEc

                    break

            #-----------------Determin incadrarea termenului urban, rural------------
            myListDay = str(delta_preg_term).split()
            try:
                if int(myListDay[0]):
                    myList = myListDay[2].split(":")
                    myHour = 24 * int(myListDay[0]) + int(myList[0])
                    myMinute = myHour * 60 + int(myList[1])
                    delta_preg_term = str(myHour) + ":" + myList[1] + ":00"
            except ValueError:
                myList = str(delta_preg_term).split(":")
                myMinute = int(myList[0]) * 60 + int(myList[1])
            termText = "Incadrat"
            compens = 0

            if myMinute > 8 * 60:
                difMinute = myMinute - 8 * 60
                termText = "Depasit cu: " + str(round(difMinute/60)) + "H " + myList[1] + "min"
                if difMinute / 60 <= 3:
                    k = 1
                elif (difMinute / 60 > 3) and (difMinute / 60 <= 6):
                    k = 4
                elif (difMinute / 60 > 6) and (difMinute / 60 <= 9):
                    k = 7
                elif (difMinute / 60 > 9):
                    k = 10
                compens = round(0.01 * 160 * 2.04 * k, 2)


            ###################### Pregatesc variabilele pentru lucru excel ######################
            # excel_control = True

            # full_file = download_file(
            #     "Deconectari.xlsx",
            #     "1KlOQQNZROaqs8C4xAQX1-kUaYNMCg9nz",
            #     SAIDI_LOCATION
            # )
            # wbDec = load_workbook(full_file)
            # wsNepr = wbDec["neprogramate-JT"]
            # wsProg = wbDec["programate-MT+JT"]



            # full_file_al = download_file(
            #     "Registru AUTORIZATIILOR UN.xlsx",
            #     "1EOclaLt1B-lCvMgYBOF6EGAaVjBBayFs",
            #     REGAL_LOCATION
            # )
            # wbAl = load_workbook(full_file_al)
            # wsAl = wbAl["Registru"]
            #-----------------Introduc datele mongodb, analiza anuala si lunara
            if cu_dec == "Programat":
                anlz_reg = f'anlzan{date_get()[5]}p'
                dec_reg = f'decpr_{date_get()[0]}_{date_get()[1]}'

                for i in mongo.db[anlz_reg].find().sort("_id", -1).limit(1):
                    nrAnlz = i["_id"]
                
                try:
                    nrAnlz = nrAnlz + 1
                except UnboundLocalError:
                    nrAnlz = 1

                for i in mongo.db[dec_reg].find().sort("_id", -1).limit(1):
                    nrDec = i["_id"]

                try:
                    nrDec = nrDec + 1
                except UnboundLocalError:
                    nrDec = 1

                mongo.db[anlz_reg].insert_one({
                    "_id": nrAnlz,
                    "oficiul": oficiul,
                    "pt_fider": pt + " " + fid,
                    "localitate": localitatea,
                    "nr_cons": fidNrCas + fidNrEc,
                    "ore": delta_preg_term,
                    "nr_dec": 1
                })

                mongo.db[dec_reg].insert_one({
                    "_id": nrDec,
                    "oficiul": oficiul,
                    "pt": pt,
                    "fid_04kv": fid,
                    "data_dec": myPreg_date_str,
                    "data_conect": myTerm_date_str,
                    "durata": delta_preg_term,
                    "cons_cas": fidNrCas,
                    "cons_ec": fidNrEc,
                    "total": fidNrCas + fidNrEc,
                    "localitate": localitatea,
                    "cauza": lucrarile,
                    "termen": termText,
                    "compens": compens,
                })

                

            #-----------------Introduc datele in postgres, analiza anuala si lunara------------
            # if cu_dec == "Programat":
            #     cur.execute("""SELECT * FROM pg_tables WHERE SCHEMANAME='public'""")
            #     isTable = False
            #     for table_tuple in cur.fetchall():
            #         if table_tuple[1] == f'anlzan{date_get()[5]}p':
            #             isTable = True
            #             break
            #     if not isTable:
            #         cur.execute(f"""CREATE TABLE anlzan{date_get()[5]}p (
            #                         anlzan{date_get()[5]}p_id serial PRIMARY KEY,
            #                         oficiul VARCHAR(32),
            #                         pt_fider VARCHAR(128),
            #                         localitate VARCHAR(128),
            #                         nr_cons NUMERIC,
            #                         ore VARCHAR(32),
            #                         nr_dec NUMERIC,
            #                         nr_regl VARCHAR(128),
            #                         compens NUMERIC
            #                     )""")
            #         conn.commit()

            #     cur.execute(f"""INSERT INTO anlzan{date_get()[5]}p (
            #                     oficiul,
            #                     pt_fider,
            #                     localitate,
            #                     nr_cons,
            #                     ore,
            #                     nr_dec
            #                 ) VALUES (
            #                     '{oficiul}',
            #                     '{pt + " " + fid}',
            #                     '{localitatea}',
            #                     '{fidNrCas + fidNrEc}',
            #                     '{delta_preg_term}',
            #                     '{1}'
            #                 )"""
            #             )
            #     conn.commit()

            #     cur.execute("""SELECT * FROM pg_tables WHERE SCHEMANAME='public'""")
            #     isTable = False
            #     for table_tuple in cur.fetchall():
            #         if table_tuple[1] == f"decpr_{date_get()[0]}_{date_get()[1]}":
            #             isTable = True
            #             break
            #     if not isTable:
            #         cur.execute(f"""CREATE TABLE decpr_{date_get()[0]}_{date_get()[1]} (
            #                                                 decpr_id serial PRIMARY KEY,
            #                                                 oficiul VARCHAR(32),
            #                                                 nr_ordine INT4,
            #                                                 pt VARCHAR(32),
            #                                                 fid_04kv VARCHAR(32),
            #                                                 data_dec VARCHAR(32),
            #                                                 data_conect VARCHAR(32),
            #                                                 durata VARCHAR(32),
            #                                                 cons_cas INT4,
            #                                                 cons_ec INT4,
            #                                                 total INT4,
            #                                                 localitate VARCHAR(32),
            #                                                 cauza VARCHAR(256),
            #                                                 termen VARCHAR(256),
            #                                                 compens NUMERIC            
            #                     )""")
            #         conn.commit()

            #     cur.execute(f"""SELECT decpr_id 
            #                     FROM decpr_{date_get()[0]}_{date_get()[1]}
            #                     ORDER BY decpr_id DESC LIMIT 1""")
            #     decpr_id_tuple = cur.fetchall()
            #     if len(decpr_id_tuple) == 0:
            #         nrDec = 1
            #     else:
            #         nrDec = decpr_id_tuple[0][0] + 1
            #     # print(cur.fetchall())
            #     # print(nrDec)
            #     cur.execute(f"""INSERT INTO decpr_{date_get()[0]}_{date_get()[1]} (
            #                         oficiul, 
            #                         nr_ordine, 
            #                         pt, 
            #                         fid_04kv, 
            #                         data_dec, 
            #                         data_conect, 
            #                         durata, 
            #                         cons_cas, 
            #                         cons_ec, 
            #                         total, 
            #                         localitate, 
            #                         cauza, 
            #                         termen, 
            #                         compens
            #                 ) VALUES (
            #                         '{oficiul}',
            #                         '{nrDec}',
            #                         '{pt}',
            #                         '{fid}',
            #                         '{myPreg_date_str}',
            #                         '{myTerm_date_str}',
            #                         '{delta_preg_term}',
            #                         '{fidNrCas}',
            #                         '{fidNrEc}',
            #                         '{fidNrCas + fidNrEc}',
            #                         '{localitatea}',
            #                         '{lucrarile}',
            #                         '{termText}',
            #                         '{compens}'
            #                 )"""
            #             )
            #     conn.commit()


                

                # max_row = wsProg.max_row + 1

                # wsProg.cell(row=max_row, column=1).value = \
                #     wsProg.cell(row=max_row - 1, column=1).value + 1
                # wsProg.cell(row=max_row, column=1).alignment = \
                #     Alignment(horizontal="center", vertical="center")
                #
                # wsProg.cell(row=max_row, column=2).value = oficiul
                # wsProg.cell(row=max_row, column=2).alignment = \
                #     Alignment(horizontal="center", vertical="center")
                #
                # wsProg.cell(row=max_row, column=3).value = \
                #     date_get()[2] + "." + date_get()[1] + "." + date_get()[5]
                # wsProg.cell(row=max_row, column=3).alignment = \
                #     Alignment(horizontal="center", vertical="center")
                #
                # wsProg.cell(row=max_row, column=4).value = \
                #     "JT"
                # wsProg.cell(row=max_row, column=4).alignment = \
                #     Alignment(horizontal="center", vertical="center")
                #
                # wsProg.cell(row=max_row, column=5).value = \
                #     pt
                # wsProg.cell(row=max_row, column=5).alignment = \
                #     Alignment(horizontal="center", vertical="center")
                #
                # wsProg.cell(row=max_row, column=6).value = \
                #     fid
                # wsProg.cell(row=max_row, column=6).alignment = \
                #     Alignment(horizontal="center", vertical="center")
                #
                # wsProg.cell(row=max_row, column=7).value = \
                #     0
                # wsProg.cell(row=max_row, column=7).alignment = \
                #     Alignment(horizontal="center", vertical="center")
                #
                # wsProg.cell(row=max_row, column=8).value = \
                #     1
                # wsProg.cell(row=max_row, column=8).alignment = \
                #     Alignment(horizontal="center", vertical="center")
                #
                # wsProg.cell(row=max_row, column=9).value = \
                #     1
                # wsProg.cell(row=max_row, column=9).alignment = \
                #     Alignment(horizontal="center", vertical="center")
                #
                # wsProg.cell(row=max_row, column=10).value = \
                #     myPreg_date_str
                # wsProg.cell(row=max_row, column=10).alignment = \
                #     Alignment(horizontal="center", vertical="center")
                #
                # wsProg.cell(row=max_row, column=11).value = \
                #     myTerm_date_str
                # wsProg.cell(row=max_row, column=11).alignment = \
                #     Alignment(horizontal="center", vertical="center")
                #
                # wsProg.cell(row=max_row, column=12).value = \
                #     delta_preg_term
                # wsProg.cell(row=max_row, column=12).alignment = \
                #     Alignment(horizontal="center", vertical="center")
                #
                # wsProg.cell(row=max_row, column=13).value = \
                #     fidNrCas
                # wsProg.cell(row=max_row, column=13).alignment = \
                #     Alignment(horizontal="center", vertical="center")
                #
                # wsProg.cell(row=max_row, column=14).value = \
                #     fidNrEc
                # wsProg.cell(row=max_row, column=14).alignment = \
                #     Alignment(horizontal="center", vertical="center")
                #
                # wsProg.cell(row=max_row, column=15).value = \
                #     fidNrCas + fidNrEc
                # wsProg.cell(row=max_row, column=15).alignment = \
                #     Alignment(horizontal="center", vertical="center")
                #
                # wsProg.cell(row=max_row, column=16).value = \
                #     localitatea
                # wsProg.cell(row=max_row, column=16).alignment = \
                #     Alignment(horizontal="center", vertical="center")
                #
                # wsProg.cell(row=max_row, column=19).value = \
                #     lucrarile
                # # wsProg.cell(row=max_row, column=19).alignment = \
                # #     Alignment(horizontal="center", vertical="center")
                #
                # try:
                #     wbDec.save(full_file)
                # except (PermissionError):
                #     excel_control = "non_excel"
                #     # return "non_excel"

            elif cu_dec == "Neprogramat":
                anlz_reg = f'anlzan{date_get()[5]}n'
                dec_reg = f'decnepr_{date_get()[0]}_{date_get()[1]}'

                for i in mongo.db[anlz_reg].find().sort("_id", -1).limit(1):
                    nrAnlz = i["_id"]

                try:
                    nrAnlz = nrAnlz + 1
                except UnboundLocalError:
                    nrAnlz = 1

                
                for i in mongo.db[dec_reg].find().sort("_id", -1).limit(1):
                    nrDec = i["_id"]

                try:
                    nrDec = nrDec + 1
                except UnboundLocalError:
                    nrDec = 1

                mongo.db[anlz_reg].insert_one({
                    "_id": nrAnlz,
                    "oficiul": oficiul,
                    "pt_fider": pt + " " + fid,
                    "localitate": localitatea,
                    "nr_cons": fidNrCas + fidNrEc,
                    "ore": delta_preg_term,
                    "nr_dec": 1
                })

                mongo.db[dec_reg].insert_one({
                    "_id": nrDec,
                    "oficiul": oficiul,
                    "pt": pt,
                    "fid_04kv": fid,
                    "data_dec": myPreg_date_str,
                    "data_conect": myTerm_date_str,
                    "durata": delta_preg_term,
                    "cons_cas": fidNrCas,
                    "cons_ec": fidNrEc,
                    "total": fidNrCas + fidNrEc,
                    "localitate": localitatea,
                    "cauza": lucrarile,
                    "termen": termText,
                    "compens": compens,
                })


            # elif cu_dec == "Neprogramat":
            #     cur.execute("""SELECT * FROM pg_tables WHERE SCHEMANAME='public'""")
            #     isTable = False
            #     for table_tuple in cur.fetchall():
            #         if table_tuple[1] == f'anlzan{date_get()[5]}n':
            #             isTable = True
            #             break
            #     if not isTable:
            #         cur.execute(f"""CREATE TABLE anlzan{date_get()[5]}n (
            #                         anlzan{date_get()[5]}n_id serial PRIMARY KEY,
            #                         oficiul VARCHAR(32),
            #                         pt_fider VARCHAR(128),
            #                         localitate VARCHAR(128),
            #                         nr_cons NUMERIC,
            #                         ore VARCHAR(32),
            #                         nr_dec NUMERIC,
            #                         nr_regl VARCHAR(128),
            #                         compens NUMERIC
            #                     )""")
            #         conn.commit()

            #     cur.execute(f"""INSERT INTO anlzan{date_get()[5]}n (
            #                     oficiul,
            #                     pt_fider,
            #                     localitate,
            #                     nr_cons,
            #                     ore,
            #                     nr_dec
            #                 ) VALUES (
            #                     '{oficiul}',
            #                     '{pt + " " + fid}',
            #                     '{localitatea}',
            #                     '{fidNrCas + fidNrEc}',
            #                     '{delta_preg_term}',
            #                     '{1}'
            #                 )"""
            #             )
            #     conn.commit()

            #     cur.execute("""SELECT * FROM pg_tables WHERE SCHEMANAME='public'""")
            #     isTable = False
            #     for table_tuple in cur.fetchall():
            #         if table_tuple[1] == f"decnepr_{date_get()[0]}_{date_get()[1]}":
            #             isTable = True
            #             break
            #     if not isTable:
            #         cur.execute(f"""CREATE TABLE decnepr_{date_get()[0]}_{date_get()[1]} (
            #                                                     decnepr_id serial PRIMARY KEY,
            #                                                     oficiul VARCHAR(32),
            #                                                     nr_ordine INT4,
            #                                                     pt VARCHAR(32),
            #                                                     fid_04kv VARCHAR(32),
            #                                                     data_dec VARCHAR(32),
            #                                                     data_conect VARCHAR(32),
            #                                                     durata VARCHAR(32),
            #                                                     cons_cas INT4,
            #                                                     cons_ec INT4,
            #                                                     total INT4,
            #                                                     localitate VARCHAR(32),
            #                                                     cauza VARCHAR(256),
            #                                                     termen VARCHAR(256),
            #                                                     compens NUMERIC            
            #                         )""")
            #         conn.commit()

            #     cur.execute(f"""SELECT decnepr_id 
            #                             FROM decnepr_{date_get()[0]}_{date_get()[1]}
            #                             ORDER BY decnepr_id DESC LIMIT 1""")
            #     decnepr_id_tuple = cur.fetchall()
            #     if len(decnepr_id_tuple) == 0:
            #         nrDec = 1
            #     else:
            #         nrDec = decnepr_id_tuple[0][0] + 1

            #     cur.execute(f"""INSERT INTO decnepr_{date_get()[0]}_{date_get()[1]} (
            #                                 oficiul, 
            #                                 nr_ordine, 
            #                                 pt, 
            #                                 fid_04kv, 
            #                                 data_dec, 
            #                                 data_conect, 
            #                                 durata, 
            #                                 cons_cas, 
            #                                 cons_ec, 
            #                                 total, 
            #                                 localitate, 
            #                                 cauza, 
            #                                 termen, 
            #                                 compens
            #                         ) VALUES (
            #                                 '{oficiul}',
            #                                 '{nrDec}',
            #                                 '{pt}',
            #                                 '{fid}',
            #                                 '{myPreg_date_str}',
            #                                 '{myTerm_date_str}',
            #                                 '{delta_preg_term}',
            #                                 '{fidNrCas}',
            #                                 '{fidNrEc}',
            #                                 '{fidNrCas + fidNrEc}',
            #                                 '{localitatea}',
            #                                 '{lucrarile}',
            #                                 '{termText}',
            #                                 '{compens}'
            #                         )"""
            #                 )
            #     conn.commit()

                # max_row = wsNepr.max_row + 1
                #
                # wsNepr.cell(row=max_row, column=1).value = \
                #     wsNepr.cell(row=max_row - 1, column=1).value + 1
                # wsNepr.cell(row=max_row, column=1).alignment = \
                #     Alignment(horizontal="center", vertical="center")
                #
                # wsNepr.cell(row=max_row, column=2).value = oficiul
                # wsNepr.cell(row=max_row, column=2).alignment = \
                #     Alignment(horizontal="center", vertical="center")
                #
                # wsNepr.cell(row=max_row, column=3).value = \
                #     date_get()[2] + "." + date_get()[1] + "." + date_get()[5]
                # wsNepr.cell(row=max_row, column=3).alignment = \
                #     Alignment(horizontal="center", vertical="center")
                #
                # wsNepr.cell(row=max_row, column=4).value = \
                #     pt
                # wsNepr.cell(row=max_row, column=4).alignment = \
                #     Alignment(horizontal="center", vertical="center")
                #
                # wsNepr.cell(row=max_row, column=5).value = \
                #     fid
                # wsNepr.cell(row=max_row, column=5).alignment = \
                #     Alignment(horizontal="center", vertical="center")
                #
                # wsNepr.cell(row=max_row, column=6).value = \
                #     0
                # wsNepr.cell(row=max_row, column=6).alignment = \
                #     Alignment(horizontal="center", vertical="center")
                #
                # wsNepr.cell(row=max_row, column=7).value = \
                #     0
                # wsNepr.cell(row=max_row, column=7).alignment = \
                #     Alignment(horizontal="center", vertical="center")
                #
                # wsNepr.cell(row=max_row, column=8).value = \
                #     1
                # wsNepr.cell(row=max_row, column=8).alignment = \
                #     Alignment(horizontal="center", vertical="center")
                #
                # wsNepr.cell(row=max_row, column=9).value = \
                #     1
                # wsNepr.cell(row=max_row, column=9).alignment = \
                #     Alignment(horizontal="center", vertical="center")
                #
                # wsNepr.cell(row=max_row, column=10).value = \
                #     myPreg_date_str
                # wsNepr.cell(row=max_row, column=10).alignment = \
                #     Alignment(horizontal="center", vertical="center")
                #
                # wsNepr.cell(row=max_row, column=11).value = \
                #     myTerm_date_str
                # wsNepr.cell(row=max_row, column=11).alignment = \
                #     Alignment(horizontal="center", vertical="center")
                #
                # wsNepr.cell(row=max_row, column=19).value = \
                #     delta_preg_term
                # wsNepr.cell(row=max_row, column=19).alignment = \
                #     Alignment(horizontal="center", vertical="center")
                #
                # wsNepr.cell(row=max_row, column=25).value = \
                #     fidNrCas
                # wsNepr.cell(row=max_row, column=25).alignment = \
                #     Alignment(horizontal="center", vertical="center")
                #
                # wsNepr.cell(row=max_row, column=26).value = \
                #     fidNrEc
                # wsNepr.cell(row=max_row, column=26).alignment = \
                #     Alignment(horizontal="center", vertical="center")
                #
                # wsNepr.cell(row=max_row, column=27).value = \
                #     fidNrCas + fidNrEc
                # wsNepr.cell(row=max_row, column=27).alignment = \
                #     Alignment(horizontal="center", vertical="center")
                #
                # wsNepr.cell(row=max_row, column=28).value = \
                #     localitatea
                # wsNepr.cell(row=max_row, column=28).alignment = \
                #     Alignment(horizontal="center", vertical="center")
                #
                # wsNepr.cell(row=max_row, column=32).value = \
                #     lucrarile
                # # wsNepr.cell(row=max_row, column=32).alignment = \
                # #     Alignment(horizontal="center", vertical="center")
                #
                # try:
                #     wbDec.save(full_file)
                # except (PermissionError):
                #     excel_control = "non_excel"
                    # return "non_excel"

            #----------------------- Introduc datele in registru autorizatii excel ----------------------------
            # max_row = wsAl.max_row + 1
            #
            # for i in mongo.db[reg_al].find({"_id": int(dsAlId)}):
            #     wsAl.cell(row=max_row, column=1).value = i["oficiul"]
            #     wsAl.cell(row=max_row, column=1).alignment = \
            #         Alignment(horizontal="center", vertical="center", wrap_text=True)
            #     wsAl.cell(row=max_row, column=2).value = i["nr_al"]
            #     wsAl.cell(row=max_row, column=2).alignment = \
            #         Alignment(horizontal="center", vertical="center", wrap_text=True)
            #     wsAl.cell(row=max_row, column=3).value = i["cu_dec"]
            #     wsAl.cell(row=max_row, column=3).alignment = \
            #         Alignment(horizontal="center", vertical="center", wrap_text=True)
            #     wsAl.cell(row=max_row, column=4).value = i["instalatia"]
            #     wsAl.cell(row=max_row, column=4).alignment = \
            #         Alignment(horizontal="center", vertical="center", wrap_text=True)
            #     wsAl.cell(row=max_row, column=5).value = i["pt"]
            #     wsAl.cell(row=max_row, column=5).alignment = \
            #         Alignment(horizontal="center", vertical="center", wrap_text=True)
            #     wsAl.cell(row=max_row, column=6).value = i["fid_nr"]
            #     wsAl.cell(row=max_row, column=6).alignment = \
            #         Alignment(horizontal="center", vertical="center", wrap_text=True)
            #     wsAl.cell(row=max_row, column=7).value = i["localitatea"]
            #     wsAl.cell(row=max_row, column=7).alignment = \
            #         Alignment(horizontal="center", vertical="center", wrap_text=True)
            #     wsAl.cell(row=max_row, column=8).value = i["lucrarile"]
            #     wsAl.cell(row=max_row, column=8).alignment = \
            #         Alignment(horizontal="center", vertical="center", wrap_text=True)
            #     wsAl.cell(row=max_row, column=9).value = i["sef"]
            #     wsAl.cell(row=max_row, column=9).alignment = \
            #         Alignment(horizontal="center", vertical="center", wrap_text=True)
            #     wsAl.cell(row=max_row, column=10).value = i["mem_ech"]
            #     wsAl.cell(row=max_row, column=10).alignment = \
            #         Alignment(horizontal="center", vertical="center", wrap_text=True)
            #     wsAl.cell(row=max_row, column=11).value = i["emitent"]
            #     wsAl.cell(row=max_row, column=11).alignment = \
            #         Alignment(horizontal="center", vertical="center", wrap_text=True)
            #     wsAl.cell(row=max_row, column=12).value = i["mas_teh"]
            #     wsAl.cell(row=max_row, column=12).alignment = \
            #         Alignment(horizontal="center", vertical="center", wrap_text=True)
            #     wsAl.cell(row=max_row, column=13).value = i["semnatura"]
            #     wsAl.cell(row=max_row, column=13).alignment = \
            #         Alignment(horizontal="center", vertical="center", wrap_text=True)
            #     wsAl.cell(row=max_row, column=14).value = i["starea"]
            #     wsAl.cell(row=max_row, column=14).alignment = \
            #         Alignment(horizontal="center", vertical="center", wrap_text=True)
            #     wsAl.cell(row=max_row, column=15).value = i["pregatire"]
            #     wsAl.cell(row=max_row, column=15).alignment = \
            #         Alignment(horizontal="center", vertical="center", wrap_text=True)
            #     wsAl.cell(row=max_row, column=16).value = i["admitere"]
            #     wsAl.cell(row=max_row, column=16).alignment = \
            #         Alignment(horizontal="center", vertical="center", wrap_text=True)
            #     wsAl.cell(row=max_row, column=17).value = i["terminare"]
            #     wsAl.cell(row=max_row, column=17).alignment = \
            #         Alignment(horizontal="center", vertical="center", wrap_text=True)
            #
            #     try:
            #         wbAl.save(full_file_al)
            #     except (PermissionError):
            #         excel_control = "non_excel_al"
            #         # return "non_excel_al"
            #
            # if excel_control != True:
            #     return excel_control

        return "approved"
    else:
        return "non-approved"

@app.route('/nepr-lunar/excel/<mth>/')
def nepr_lunar_excel(mth):
    if mth == "undefined":
        mth = date_get()[1]

    full_file = download_file(
        "Deconectari.xlsx",
        "1KlOQQNZROaqs8C4xAQX1-kUaYNMCg9nz",
        SAIDI_LOCATION,
        mth
    )
    wbDec = load_workbook(full_file)
    wsNepr = wbDec["neprogramate-JT"]

    max_row = wsNepr.max_row

    dec_reg = f"decnepr_{date_get()[0]}_{mth}"

    j = 1
    # cur.execute(f"SELECT * FROM decnepr_{date_get()[0]}_{mth}")
    # for i in cur.fetchall():
    for i in mongo.db[dec_reg].find():
        if i["_id"] > wsNepr.cell(row=max_row, column=1).value:
            myDate = i["data_dec"].split()
            myDate = myDate[0]

            wsNepr.cell(row=max_row + j, column=1).value = i["_id"]
            wsNepr.cell(row=max_row + j, column=1).alignment = \
                Alignment(horizontal="center", vertical="center")

            wsNepr.cell(row=max_row + j, column=2).value = i["oficiul"]
            wsNepr.cell(row=max_row + j, column=2).alignment = \
                Alignment(horizontal="center", vertical="center")

            wsNepr.cell(row=max_row + j, column=3).value = myDate
            wsNepr.cell(row=max_row + j, column=3).alignment = \
                Alignment(horizontal="center", vertical="center")

            wsNepr.cell(row=max_row + j, column=4).value = i["pt"]
            wsNepr.cell(row=max_row + j, column=4).alignment = \
                Alignment(horizontal="center", vertical="center")

            wsNepr.cell(row=max_row + j, column=5).value = i["fid_04kv"]
            wsNepr.cell(row=max_row + j, column=5).alignment = \
                Alignment(horizontal="center", vertical="center")

            wsNepr.cell(row=max_row + j, column=6).value = \
                0
            wsNepr.cell(row=max_row + j, column=6).alignment = \
                Alignment(horizontal="center", vertical="center")

            wsNepr.cell(row=max_row + j, column=7).value = \
                0
            wsNepr.cell(row=max_row + j, column=7).alignment = \
                Alignment(horizontal="center", vertical="center")

            wsNepr.cell(row=max_row + j, column=8).value = \
                1
            wsNepr.cell(row=max_row + j, column=8).alignment = \
                Alignment(horizontal="center", vertical="center")

            wsNepr.cell(row=max_row + j, column=9).value = \
                1
            wsNepr.cell(row=max_row + j, column=9).alignment = \
                Alignment(horizontal="center", vertical="center")

            wsNepr.cell(row=max_row + j, column=10).value = i["data_dec"]
            wsNepr.cell(row=max_row + j, column=10).alignment = \
                Alignment(horizontal="center", vertical="center")

            wsNepr.cell(row=max_row + j, column=11).value = i["data_conect"]
            wsNepr.cell(row=max_row + j, column=11).alignment = \
                Alignment(horizontal="center", vertical="center")

            wsNepr.cell(row=max_row + j, column=19).value = i["durata"]
            wsNepr.cell(row=max_row + j, column=19).alignment = \
                Alignment(horizontal="center", vertical="center")

            wsNepr.cell(row=max_row + j, column=25).value = i["cons_cas"]
            wsNepr.cell(row=max_row + j, column=25).alignment = \
                Alignment(horizontal="center", vertical="center")

            wsNepr.cell(row=max_row + j, column=26).value = i["cons_ec"]
            wsNepr.cell(row=max_row + j, column=26).alignment = \
                Alignment(horizontal="center", vertical="center")

            wsNepr.cell(row=max_row + j, column=27).value = i["total"]
            wsNepr.cell(row=max_row + j, column=27).alignment = \
                Alignment(horizontal="center", vertical="center")

            wsNepr.cell(row=max_row + j, column=28).value = i["localitate"]
            wsNepr.cell(row=max_row + j, column=28).alignment = \
                Alignment(horizontal="center", vertical="center")

            wsNepr.cell(row=max_row + j, column=32).value = i["cauza"]
            # wsNepr.cell(row=max_row, column=32).alignment = \
            #     Alignment(horizontal="center", vertical="center")
            j += 1

    try:
        wbDec.save(full_file)
    except (PermissionError):
        return "nonexcel", 403

    return "ok", 200

@app.route('/prog-lunar/excel/<mth>/')
def prog_lunar_excel(mth):
    if mth == "undefined":
        mth = date_get()[1]

    full_file = download_file(
        "Deconectari.xlsx",
        "1KlOQQNZROaqs8C4xAQX1-kUaYNMCg9nz",
        SAIDI_LOCATION,
        mth
    )
    wbDec = load_workbook(full_file)
    wsProg = wbDec["programate-MT+JT"]

    max_row = wsProg.max_row
    # print(wsProg.cell(row=max_row, column=1).value)
    j = 1
    cur.execute(f"SELECT * FROM decpr_{date_get()[0]}_{mth}")
    for i in cur.fetchall():
        if i[0] > wsProg.cell(row=max_row, column=1).value:
            myDate = i[5].split()
            myDate = myDate[0]

            wsProg.cell(row=max_row + j, column=1).value = i[0]
            wsProg.cell(row=max_row + j, column=1).alignment = \
                Alignment(horizontal="center", vertical="center")

            wsProg.cell(row=max_row + j, column=2).value = i[1]
            wsProg.cell(row=max_row + j, column=2).alignment = \
                Alignment(horizontal="center", vertical="center")

            wsProg.cell(row=max_row + j, column=3).value = myDate
            wsProg.cell(row=max_row + j, column=3).alignment = \
                Alignment(horizontal="center", vertical="center")

            wsProg.cell(row=max_row + j, column=4).value = "JT"
            wsProg.cell(row=max_row + j, column=4).alignment = \
                Alignment(horizontal="center", vertical="center")

            wsProg.cell(row=max_row + j, column=5).value = i[3]
            wsProg.cell(row=max_row + j, column=5).alignment = \
                Alignment(horizontal="center", vertical="center")

            wsProg.cell(row=max_row + j, column=6).value = i[4]
            wsProg.cell(row=max_row + j, column=6).alignment = \
                Alignment(horizontal="center", vertical="center")

            wsProg.cell(row=max_row + j, column=7).value = \
                0
            wsProg.cell(row=max_row + j, column=7).alignment = \
                Alignment(horizontal="center", vertical="center")

            wsProg.cell(row=max_row + j, column=8).value = \
                1
            wsProg.cell(row=max_row + j, column=8).alignment = \
                Alignment(horizontal="center", vertical="center")

            wsProg.cell(row=max_row + j, column=9).value = \
                1
            wsProg.cell(row=max_row + j, column=9).alignment = \
                Alignment(horizontal="center", vertical="center")

            wsProg.cell(row=max_row + j, column=10).value = i[5]
            wsProg.cell(row=max_row + j, column=10).alignment = \
                Alignment(horizontal="center", vertical="center")

            wsProg.cell(row=max_row + j, column=11).value = i[6]
            wsProg.cell(row=max_row + j, column=11).alignment = \
                Alignment(horizontal="center", vertical="center")

            wsProg.cell(row=max_row + j, column=12).value = i[7]
            wsProg.cell(row=max_row + j, column=12).alignment = \
                Alignment(horizontal="center", vertical="center")

            wsProg.cell(row=max_row + j, column=13).value = i[8]
            wsProg.cell(row=max_row + j, column=13).alignment = \
                Alignment(horizontal="center", vertical="center")

            wsProg.cell(row=max_row + j, column=14).value = i[9]
            wsProg.cell(row=max_row + j, column=14).alignment = \
                Alignment(horizontal="center", vertical="center")

            wsProg.cell(row=max_row + j, column=15).value = i[10]
            wsProg.cell(row=max_row + j, column=15).alignment = \
                Alignment(horizontal="center", vertical="center")

            wsProg.cell(row=max_row + j, column=16).value = i[11]
            wsProg.cell(row=max_row + j, column=16).alignment = \
                Alignment(horizontal="center", vertical="center")

            wsProg.cell(row=max_row + j, column=19).value = i[12]
            # wsProg.cell(row=max_row, column=19).alignment = \
            #     Alignment(horizontal="center", vertical="center")
            j += 1

    try:
        wbDec.save(full_file)
    except (PermissionError):
        return "nonexcel", 403

    return "ok", 200


@app.route('/autorizatii/excel/<mth>/')
def exp_excel(mth):
    if mth == "undefined":
        mth = date_get()[1]

    full_file_al = download_file(
        "Registru AUTORIZATIILOR UN.xlsx",
        "1EOclaLt1B-lCvMgYBOF6EGAaVjBBayFs",
        REGAL_LOCATION,
        mth
    )
    wbAl = load_workbook(full_file_al)
    wsAl = wbAl["Registru"]

    max_row = wsAl.max_row
    # print(wsAl.cell(row=max_row, column=2).value)

    reg_al = "al_" + date_get()[0] + "_" + mth
    j = 1
    for i in mongo.db[reg_al].find():
        # print(wsAl.cell(row=max_row, column=2).value)
        if i['_id'] > wsAl.cell(row=max_row, column=2).value and \
                i['terminare'] != 'Terminare':
            wsAl.cell(row=max_row + j, column=1).value = i["oficiul"]
            wsAl.cell(row=max_row + j, column=1).alignment = \
                Alignment(horizontal="center", vertical="center", wrap_text=True)
            wsAl.cell(row=max_row + j, column=2).value = i["nr_al"]
            wsAl.cell(row=max_row + j, column=2).alignment = \
                Alignment(horizontal="center", vertical="center", wrap_text=True)
            wsAl.cell(row=max_row + j, column=3).value = i["cu_dec"]
            wsAl.cell(row=max_row + j, column=3).alignment = \
                Alignment(horizontal="center", vertical="center", wrap_text=True)
            wsAl.cell(row=max_row + j, column=4).value = i["instalatia"]
            wsAl.cell(row=max_row + j, column=4).alignment = \
                Alignment(horizontal="center", vertical="center", wrap_text=True)
            wsAl.cell(row=max_row + j, column=5).value = i["pt"]
            wsAl.cell(row=max_row + j, column=5).alignment = \
                Alignment(horizontal="center", vertical="center", wrap_text=True)
            wsAl.cell(row=max_row + j, column=6).value = i["fid_nr"]
            wsAl.cell(row=max_row + j, column=6).alignment = \
                Alignment(horizontal="center", vertical="center", wrap_text=True)
            wsAl.cell(row=max_row + j, column=7).value = i["localitatea"]
            wsAl.cell(row=max_row + j, column=7).alignment = \
                Alignment(horizontal="center", vertical="center", wrap_text=True)
            wsAl.cell(row=max_row + j, column=8).value = i["lucrarile"]
            wsAl.cell(row=max_row + j, column=8).alignment = \
                Alignment(horizontal="center", vertical="center", wrap_text=True)
            wsAl.cell(row=max_row + j, column=9).value = i["sef"]
            wsAl.cell(row=max_row + j, column=9).alignment = \
                Alignment(horizontal="center", vertical="center", wrap_text=True)
            wsAl.cell(row=max_row + j, column=10).value = i["mem_ech"]
            wsAl.cell(row=max_row + j, column=10).alignment = \
                Alignment(horizontal="center", vertical="center", wrap_text=True)
            wsAl.cell(row=max_row + j, column=11).value = i["emitent"]
            wsAl.cell(row=max_row + j, column=11).alignment = \
                Alignment(horizontal="center", vertical="center", wrap_text=True)
            wsAl.cell(row=max_row + j, column=12).value = i["mas_teh"]
            wsAl.cell(row=max_row + j, column=12).alignment = \
                Alignment(horizontal="center", vertical="center", wrap_text=True)
            wsAl.cell(row=max_row + j, column=13).value = i["semnatura"]
            wsAl.cell(row=max_row + j, column=13).alignment = \
                Alignment(horizontal="center", vertical="center", wrap_text=True)
            wsAl.cell(row=max_row + j, column=14).value = i["starea"]
            wsAl.cell(row=max_row + j, column=14).alignment = \
                Alignment(horizontal="center", vertical="center", wrap_text=True)
            wsAl.cell(row=max_row + j, column=15).value = i["pregatire"]
            wsAl.cell(row=max_row + j, column=15).alignment = \
                Alignment(horizontal="center", vertical="center", wrap_text=True)
            wsAl.cell(row=max_row + j, column=16).value = i["admitere"]
            wsAl.cell(row=max_row + j, column=16).alignment = \
                Alignment(horizontal="center", vertical="center", wrap_text=True)
            wsAl.cell(row=max_row + j, column=17).value = i["terminare"]
            wsAl.cell(row=max_row + j, column=17).alignment = \
                Alignment(horizontal="center", vertical="center", wrap_text=True)

            j += 1

    try:
        wbAl.save(full_file_al)
    except (PermissionError):
        return "nonexcel", 403

    return "ok", 200

@app.route('/autorizatii/intrerupere/<dsAlId>/', methods=['PUT'])
def get_intr(dsAlId):
    starea_list = request.get_json()["starea"].split(".")
    reg_al = "al_" + date_get()[0] + "_" + starea_list[1]
    if mongo.db[reg_al].find_one({"_id": int(dsAlId)})["terminare"] == "Terminare":
        mongo.db[reg_al].update_one({"_id": int(dsAlId)}, {
            "$set": {"terminare": request.get_json()["intrerupere"]}})
        return "approved"
    else:
        return "non-approved"

@app.route('/autorizatii/nulucrat/<dsAlId>/', methods=['PUT'])
def get_nulucr(dsAlId):
    starea_list = request.get_json()["starea"].split(".")
    reg_al = "al_" + date_get()[0] + "_" + starea_list[1]
    if mongo.db[reg_al].find_one({"_id": int(dsAlId)})["pregatire"] == "Pregatire":
        mongo.db[reg_al].update_one({"_id": int(dsAlId)}, {
            "$set": {
                    "pregatire": request.get_json()["pregatire"],
                    "admitere": request.get_json()["admitere"],
                    "terminare": request.get_json()["terminare"]
            }})
        return "approved"
    else:
        return "non-approved"

# @app.route('/autorizatii/<alId>/', methods=['GET', 'PUT'])
# def get_one_al(alId):
#     if request.method == 'PUT':
#         if request.get_json() == {"pregatire": "Cerere la pregatire"}:
#             mongo.db.al_2021_04.update_one({"_id": int(alId)}, {
#                 "$set": {"pregatire": request.get_json()["pregatire"]}
#             })
#         if request.get_json() == {"admitere": "Cerere la admitere"}:
#             mongo.db.al_2021_04.update_one({"_id": int(alId)}, {
#                 "$set": {"admitere": request.get_json()["admitere"]}
#             })
#         if request.get_json() == {"terminare": "Cerere la terminare"}:
#             mongo.db.al_2021_04.update_one({"_id": int(alId)}, {
#                 "$set": {"terminare": request.get_json()["terminare"]}
#             })
#
#     al = mongo.db.al_2021_04.find_one({"_id": int(alId)})
#     output = {
#         "_id": al["_id"],
#         "oficiul": al["oficiul"],
#         "nr_ds": al["nr_ds"],
#         "nr_al": al["nr_al"],
#         "instalatia": al["instalatia"],
#         "pt": al["pt"],
#         "localitatea": al["localitatea"],
#         "fid_nr": al["fid_nr"],
#         "lucrarile": al["lucrarile"],
#         "sef": al["sef"],
#         "mem_ech": al["mem_ech"],
#         "emitent": al["emitent"],
#         "cu_dec": al["cu_dec"],
#         "mas_teh": al["mas_teh"],
#         "semnatura": al["semnatura"],
#         "starea": al["starea"],
#         "pregatire": al["pregatire"],
#         "admitere": al["admitere"],
#         "terminare": al["terminare"],
#         "link": al["link"]
#     }
#     return output

@app.route('/autorizatii/corect/<dsAlId>/<month>/')
def get_data_corect(dsAlId, month):
    reg_al = "al_" + date_get()[0] + "_" + month
    output = mongo.db[reg_al].find_one({"_id": int(dsAlId)})

    return jsonify(output)

@app.route('/autorizatii/delete/<dsAlId>/')
def get_delete(dsAlId):
    reg_al = "al_" + date_get()[0] + "_" + date_get()[1]
    if mongo.db[reg_al].find_one({"_id": int(dsAlId)})["pregatire"] == "Pregatire":
        mongo.db[reg_al].delete_one({"_id": int(dsAlId)})

        return "approved"
    else:
        return "non-approved"

# @app.route('/autorizatii/upload-link/<dsAlId>/')
# def up_link(dsAlId):
#     reg_al = "al_" + date_get()[0] + "_" + date_get()[1]
#     full_file = mongo.db[reg_al].find_one({"_id": int(dsAlId)})["link"]
#     fileControl = os.path.isfile(full_file)
#     if not re.search("http", full_file) and full_file != "" and fileControl:
#         file_name = dsAlId + os.path.splitext(full_file)[1]
#         googleLink = "https://drive.google.com/file/d/" + goog_upload(full_file, file_name)
#         mongo.db[reg_al].update_one({"_id": int(dsAlId)}, {"$set": {"link": googleLink}})
#         return "approved"
#     else:
#         return "non-approved"

@app.route('/deranjamente/<abrOf>/', methods=['GET'])
def get_all_der(abrOf):
    output = [{
        "_id": "",
        "oficiul": "",
        "nr_ordine": "",
        "transmis": "",
        "sector": "",
        "instalatia": "",
        "fid_10kv": "",
        "pt": "",
        "fid_04kv": "",
        "continutul": "",
        "data": "",
        "responsabil": "",
        "starea": "",
    }]

    reg_deranj = "deranj_" + date_get()[0]
    if abrOf == "AD":
        for i in mongo.db[reg_deranj].find():
            output.append(i)
    else:
        for i in mongo.db[reg_deranj].find({"oficiul": abrOf}):
            output.append(i)
    # print(output)
    return jsonify(output)

# @app.route('/deranjamente/<abrOf>/')
# def get_der_of(abrOf):
#     output = []
#     if abrOf == "undefined":
#         for i in mongo.db.deranj_2021.find():
#             output.append(i)
#     else:
#         for i in mongo.db.deranj_2021.find({"oficiul": abrOf}):
#             output.append(i)
#
#     return jsonify(output)

# @app.route('/deranjamente/<abrOf>/<sectOf>/')
# def get_deranj_of(abrOf, sectOf):
#     output = []
#     # sect = "sectoare_" + abrOf.lower()
#     if abrOf != "undefined" and sectOf == "undefined":
#         for i in mongo.db.deranj_2021.find({"oficiul": abrOf}):
#             output.append(i)
#     elif abrOf == "undefined" and sectOf == "undefined":
#         for i in mongo.db.deranj_2021.find():
#             output.append(i)
#     else:
#         for i in mongo.db.deranj_2021.find({"oficiul": abrOf, "sector": sectOf}):
#             output.append(i)
#     return jsonify(output)

@app.route('/deranjamente/<abrOf>/<sect>/<sectOf>/<neex>/')
def get_deranj_neexec(abrOf, sect, sectOf, neex):
    output = [{
        "_id": "",
        "oficiul": "",
        "nr_ordine": "",
        "transmis": "",
        "sector": "",
        "instalatia": "",
        "fid_10kv": "",
        "pt": "",
        "fid_04kv": "",
        "continutul": "",
        "data": "",
        "responsabil": "",
        "starea": "",
    }]
    reg_deranj = "deranj_" + date_get()[0]
    if neex == "false":
        if abrOf == "AD":
            for i in mongo.db[reg_deranj].find():
                output.append(i)
        elif abrOf != "undefined" and sect != "undefined":
            if abrOf == sectOf:
                for i in mongo.db[reg_deranj].find({"oficiul": abrOf, "sector": sect}):
                    output.append(i)
            else:
                for i in mongo.db[reg_deranj].find({"oficiul": abrOf}):
                    output.append(i)
        else:
            for i in mongo.db[reg_deranj].find({"oficiul": abrOf}):
                output.append(i)
    else:
        if abrOf == "AD":
            for i in mongo.db[reg_deranj].find({"starea": "Neexecutat"}):
                output.append(i)
        elif abrOf != "undefined" and sect != "undefined":
            if abrOf == sectOf:
                for i in mongo.db[reg_deranj].find({"oficiul": abrOf, "sector": sect, "starea": "Neexecutat"}):
                    output.append(i)
            else:
                for i in mongo.db[reg_deranj].find({"oficiul": abrOf, "starea": "Neexecutat"}):
                    output.append(i)
        else:
            for i in mongo.db[reg_deranj].find({"oficiul": abrOf, "starea": "Neexecutat"}):
                output.append(i)

    return jsonify(output)


@app.route('/nepr-anual/<abrOf>/', methods=['GET'])
def get_nepr_anual(abrOf):
    output = [{
        "_id": "",
        "oficiul": "",
        "pt_fider": "",
        "localitate": "",
        "nr_cons": "",
        "ore": "",
        "nr_dec": "",
        "nr_regl": "",
        "compens": "",
    }]
    # output = json.dumps(output)

    dec_reg = f"anlzan{date_get()[5]}n"

    if abrOf != "AD":
        for i in mongo.db[dec_reg].find({"oficiul": abrOf}):
            output.append(i)
    else:
        for i in mongo.db[dec_reg].find():
            output.append(i)

    data = pd.DataFrame.from_dict(output)

    # cur.execute("SELECT * FROM pg_tables WHERE SCHEMANAME='public'")
    # for table_tuple in cur.fetchall():
    #     if table_tuple[1] == f"anlzan{date_get()[5]}n":
    #         if abrOf == "AD":
    #             cur.execute(f"SELECT * FROM anlzan{date_get()[5]}n")
    #         else:
    #             cur.execute(f"SELECT * FROM anlzan{date_get()[5]}n WHERE oficiul = '{abrOf}'")
    #         tuples = cur.fetchall()
    #         column_names = [
    #             "_id",
    #             "oficiul",
    #             "pt_fider",
    #             "localitate",
    #             "nr_cons",
    #             "ore",
    #             "nr_dec",
    #             "nr_regl",
    #             "compens",
    #         ]

    #         data = pd.DataFrame(tuples, columns=column_names)
    data.sort_values(by="pt_fider", inplace=True, ignore_index=True)

    if data.shape[0] > 1:
        data.drop(0, inplace=True)
        data.reset_index(inplace=True)

        for i in range(1, len(data)):
            if data.at[i, "pt_fider"] == data.at[i-1, "pt_fider"]:
                data.at[i, "nr_dec"] = data.at[i, "nr_dec"] + data.at[i-1, "nr_dec"]

                # Calculez timpul total de deconectare
                myTm_list = data.at[i, "ore"].split(":")
                myTm_list_min = int(myTm_list[0]) * 60 + int(myTm_list[1])

                myTm_list_minus = data.at[i - 1, "ore"].split(":")
                myTm_list_minus_min = int(myTm_list_minus[0]) * 60 + int(myTm_list_minus[1])

                myTm_tot_min = myTm_list_min + myTm_list_minus_min
                if myTm_tot_min < 10:
                    myColon = ":0"
                else:
                    myColon = ":"
                myTime_format = str(myTm_tot_min // 60) + myColon + str(myTm_tot_min % 60) + \
                                ":00"
                data.at[i, "ore"] = myTime_format

                data.drop(i - 1, inplace=True)
        data.sort_values(by="nr_dec", inplace=True, ignore_index=True, ascending=False)

        # Calculez termenul reglementat urban, rural, suma compensatiilor
        for i in range(len(data)):
            if data.at[i, "localitate"] == None:
                data.at[i, "localitate"] = ""
            if data.at[i, "compens"] == None:
                data.at[i, "compens"] = 0
            if data.at[i, "localitate"] == "" or bool(re.search("or[.]", data.at[i, "localitate"])):
                if data.at[i, "nr_dec"] <= 9:
                    data.at[i, "nr_regl"] = "Incadrat"
                elif data.at[i, "nr_dec"] > 9:
                    data.at[i, "nr_regl"] = "Depasit cu " + \
                                            str(data.at[i, "nr_dec"] - 9) + "dec."
                    data.at[i, "compens"] = round(0.01 * (160 * 12) * 2.04 * \
                                                    (data.at[i, "nr_dec"] - 9), 2)
                else:
                    data.at[i, "nr_regl"] = "Eroare"
            else:
                if data.at[i, "nr_dec"] <= 12:
                    data.at[i, "nr_regl"] = "Incadrat"
                elif data.at[i, "nr_dec"] > 12:
                    data.at[i, "nr_regl"] = "Depasit cu " + \
                                            str(data.at[i, "nr_dec"] - 12) + "dec."
                    data.at[i, "compens"] = round(0.01 * (160 * 12) * 2.04 * \
                                                    float((data.at[i, "nr_dec"] - 12)), 2)
                else:
                    data.at[i, "nr_regl"] = "Eroare"

        # cur.close()
    output = data.to_json(orient='records')

    return output
@app.route('/prog-anual/<abrOf>/', methods=['GET'])
def get_prog_anual(abrOf):
    output = [{
        "_id": "",
        "oficiul": "",
        "pt_fider": "",
        "localitate": "",
        "nr_cons": "",
        "ore": "",
        "nr_dec": "",
        "nr_regl": "",
        "compens": "",
    }]
    # output = json.dumps(output)

    dec_reg = f"anlzan{date_get()[5]}p"

    if abrOf != "AD":
        for i in mongo.db[dec_reg].find({"oficiul": abrOf}):
            output.append(i)
    else:
        for i in mongo.db[dec_reg].find():
            output.append(i)

    data = pd.DataFrame.from_dict(output)

    # cur.execute("SELECT * FROM pg_tables WHERE SCHEMANAME='public'")
    # for table_tuple in cur.fetchall():
    #     if table_tuple[1] == f"anlzan{date_get()[5]}p":
    #         if abrOf == "AD":
    #             cur.execute(f"SELECT * FROM anlzan{date_get()[5]}p")
    #         else:
    #             cur.execute(f"SELECT * FROM anlzan{date_get()[5]}p WHERE oficiul = '{abrOf}'")
    #         tuples = cur.fetchall()
    #         column_names = [
    #             "_id",
    #             "oficiul",
    #             "pt_fider",
    #             "localitate",
    #             "nr_cons",
    #             "ore",
    #             "nr_dec",
    #             "nr_regl",
    #             "compens",
    #         ]
    #         print(tuples)
    #         data = pd.DataFrame(tuples, columns=column_names)
    data.sort_values(by="pt_fider", inplace=True, ignore_index=True)

    if data.shape[0] > 1:
        data.drop(0, inplace=True)
        data.reset_index(inplace=True)

        for i in range(1, len(data)):
            if data.at[i, "pt_fider"] == data.at[i-1, "pt_fider"]:
                data.at[i, "nr_dec"] = data.at[i, "nr_dec"] + data.at[i-1, "nr_dec"]

                # Calculez timpul total de deconectare
                myTm_list = str(data.at[i, "ore"]).split(":")
                myTm_list_min = int(myTm_list[0]) * 60 + int(myTm_list[1])

                myTm_list_minus = data.at[i - 1, "ore"].split(":")
                myTm_list_minus_min = int(myTm_list_minus[0]) * 60 + int(myTm_list_minus[1])

                myTm_tot_min = myTm_list_min + myTm_list_minus_min
                if myTm_tot_min < 10:
                    myColon = ":0"
                else:
                    myColon = ":"
                myTime_format = str(myTm_tot_min // 60) + myColon + str(myTm_tot_min % 60) + \
                                ":00"
                data.at[i, "ore"] = myTime_format

                data.drop(i - 1, inplace=True)

        data.sort_values(by="nr_dec", inplace=True, ignore_index=True, ascending=False)
        # Calculez termenul reglementat urban, rural, suma compensatiilor
        for i in range(len(data)):
            if data.at[i, "localitate"] == None:
                data.at[i, "localitate"] = ""
            if data.at[i, "compens"] == None:
                data.at[i, "compens"] = 0
            if data.at[i, "localitate"] == "" or bool(re.search("or[.]", data.at[i, "localitate"])):
                if data.at[i, "nr_dec"] <= 9:
                    data.at[i, "nr_regl"] = "Incadrat"
                elif data.at[i, "nr_dec"] > 9:
                    data.at[i, "nr_regl"] = "Depasit cu " + \
                                            str(data.at[i, "nr_dec"] - 9) + "dec."
                    data.at[i, "compens"] = round(0.01 * (160 * 12) * 2.04 * \
                                                    (data.at[i, "nr_dec"] - 9), 2)
                else:
                    data.at[i, "nr_regl"] = "Eroare"
            else:
                if data.at[i, "nr_dec"] <= 12:
                    data.at[i, "nr_regl"] = "Incadrat"
                elif data.at[i, "nr_dec"] > 12:
                    data.at[i, "nr_regl"] = "Depasit cu " + \
                                            str(data.at[i, "nr_dec"] - 12) + "dec."
                    data.at[i, "compens"] = round(0.01 * (160 * 12) * 2.04 * \
                                                    float((data.at[i, "nr_dec"] - 12)), 2)
                else:
                    data.at[i, "nr_regl"] = "Eroare"

        # cur.close()
    output = data.to_json(orient='records')

    return output

# @app.route('/nepr-lunar/', methods=['GET'])
# def get_nepr_lunar():
#     reg_dec = "decnepr_" + date_get()[0] + "_" + date_get()[1]
#     cur.execute("SELECT * FROM pg_tables WHERE SCHEMANAME='public'")
#     for table_tuple in cur.fetchall():
#         if table_tuple[1] == reg_dec:
#             cur.execute(f"SELECT * FROM {reg_dec}")
#             tuples = cur.fetchall()
#             column_names = [
#                 "_id",
#                 "oficiul",
#                 "nr_ordine",
#                 "pt",
#                 "fid_04kv",
#                 "data_dec",
#                 "data_conect",
#                 "durata",
#                 "cons_cas",
#                 "cons_ec",
#                 "total",
#                 "localitate",
#                 "cauza",
#                 "termen",
#                 "compens",
#             ]
#             data = pd.DataFrame(tuples, columns=column_names)
#             data.sort_values(by="_id", inplace=True, ignore_index=True)
#             output = data.to_json(orient='records')
#             break
#         else:
#             output = ""
#
#     return output

@app.route('/nepr-lunar/<abrOf>/<neprMth>/')
def get_nepr_mth(abrOf, neprMth):
    if neprMth == "undefined":
        neprMth = date_get()[1]

    reg_dec = "decnepr_" + date_get()[0] + "_" + neprMth

    output = [{
        "_id": "",
        "oficiul": "",
        "nr_ordine": "",
        "pt": "",
        "fid_04kv": "",
        "data_dec": "",
        "data_conect": "",
        "durata": "",
        "cons_cas": "",
        "cons_ec": "",
        "total": "",
        "localitate": "",
        "cauza": "",
        "termen": "",
        "compens": "",
    }]

    if abrOf != "AD":
        for i in mongo.db[reg_dec].find({"oficiul": abrOf}):
            output.append(i)
    else:
        for i in mongo.db[reg_dec].find():
            output.append(i)

    return jsonify(output)
    # output = json.dumps(output)

    # cur.execute("SELECT * FROM pg_tables WHERE SCHEMANAME='public'")
    # for table_tuple in cur.fetchall():
    #     if table_tuple[1] == reg_dec:
    #         if abrOf == "AD":
    #             cur.execute(f"SELECT * FROM {reg_dec}")
    #         else:
    #             cur.execute(f"SELECT * FROM {reg_dec} WHERE oficiul = '{abrOf}'")
    #         tuples = cur.fetchall()
    #         column_names = [
    #             "_id",
    #             "oficiul",
    #             "nr_ordine",
    #             "pt",
    #             "fid_04kv",
    #             "data_dec",
    #             "data_conect",
    #             "durata",
    #             "cons_cas",
    #             "cons_ec",
    #             "total",
    #             "localitate",
    #             "cauza",
    #             "termen",
    #             "compens",
    #         ]
    #         data = pd.DataFrame(tuples, columns=column_names)
    #         data.sort_values(by="_id", inplace=True, ignore_index=True)
    #         if data.shape[0] != 0:
    #             output = data.to_json(orient='records')
    #         break
    # return output

# @app.route('/prog-lunar/', methods=['GET'])
# def get_prog_lunar():
#     reg_dec = "decpr_" + date_get()[0] + "_" + date_get()[1]
#     cur.execute("SELECT * FROM pg_tables WHERE SCHEMANAME='public'")
#     for table_tuple in cur.fetchall():
#         if table_tuple[1] == reg_dec:
#             cur.execute(f"SELECT * FROM {reg_dec}")
#             tuples = cur.fetchall()
#             column_names = [
#                 "_id",
#                 "oficiul",
#                 "nr_ordine",
#                 "pt",
#                 "fid_04kv",
#                 "data_dec",
#                 "data_conect",
#                 "durata",
#                 "cons_cas",
#                 "cons_ec",
#                 "total",
#                 "localitate",
#                 "cauza",
#                 "termen",
#                 "compens",
#             ]
#             data = pd.DataFrame(tuples, columns=column_names)
#             data.sort_values(by="_id", inplace=True, ignore_index=True)
#             output = data.to_json(orient='records')
#             break
#         else:
#             output = ""
#     # print(output)
#
#     return output

@app.route('/prog-lunar/<abrOf>/<progMth>/')
def get_prog_mth(abrOf, progMth):
    if progMth == "undefined":
        progMth = date_get()[1]

    reg_dec = "decpr_" + date_get()[0] + "_" + progMth

    output = [{
        "_id": "",
        "oficiul": "",
        "nr_ordine": "",
        "pt": "",
        "fid_04kv": "",
        "data_dec": "",
        "data_conect": "",
        "durata": "",
        "cons_cas": "",
        "cons_ec": "",
        "total": "",
        "localitate": "",
        "cauza": "",
        "termen": "",
        "compens": "",
    }]

    if abrOf != "AD":
        for i in mongo.db[reg_dec].find({"oficiul": abrOf}):
            output.append(i)
    else:
        for i in mongo.db[reg_dec].find():
            output.append(i)

    return jsonify(output)
    # output = json.dumps(output)

    # cur.execute("SELECT * FROM pg_tables WHERE SCHEMANAME='public'")
    # for table_tuple in cur.fetchall():
    #     if table_tuple[1] == reg_dec:
    #         if abrOf == "AD":
    #             cur.execute(f"SELECT * FROM {reg_dec}")
    #         else:
    #             cur.execute(f"SELECT * FROM {reg_dec} WHERE oficiul = '{abrOf}'")
    #         tuples = cur.fetchall()
    #         column_names = [
    #             "_id",
    #             "oficiul",
    #             "nr_ordine",
    #             "pt",
    #             "fid_04kv",
    #             "data_dec",
    #             "data_conect",
    #             "durata",
    #             "cons_cas",
    #             "cons_ec",
    #             "total",
    #             "localitate",
    #             "cauza",
    #             "termen",
    #             "compens",
    #         ]
    #         data = pd.DataFrame(tuples, columns=column_names)
    #         data.sort_values(by="_id", inplace=True, ignore_index=True)
    #         if data.shape[0] != 0:
    #             output = data.to_json(orient='records')
    #         break
    # return output

@app.route('/oficii/')
def get_of():
    output = []
    for i in mgGen.db.oficii.find():
        output.append(i)

    output = sorted(output, key=lambda k: k['name'])
    return jsonify(output)

@app.route('/sectoare/<abrOf>/')
def get_sect(abrOf):
    output = []
    sect = "sectoare_" + abrOf.lower()
    if abrOf != "undefined":
        for i in mgGen.db[sect].find():
            output.append(i)

    output = sorted(output, key=lambda k: k['name'])
    return jsonify(output)

@app.route('/position/')
def get_position():
    output = []
    for i in mgGen.db.position.find():
        output.append(i)

    output = sorted(output, key=lambda k: k['name'])
    return jsonify(output)

@app.route('/angajati/<abrOf>/')
def get_ang_of(abrOf):
    output = []
    angaj = "angajati_" + abrOf.lower()
    for i in mgGen.db[angaj].find().sort("name", 1):
        output.append(i)

    # output = sorted(output, key=lambda k: k['name'])
    return jsonify(output)

@app.route('/log/<abrOf>/<angajName>/<enteredNr>/')
def get_log(abrOf, angajName, enteredNr):
    angaj = "angajati_" + abrOf.lower()
    if enteredNr == mgGen.db[angaj].find_one({"name": angajName})["nr_tabel"]:
        output = {
            "status": "approved",
            "position": mgGen.db[angaj].find_one({"name": angajName})["position"]
        }
    else:
        output = {
            "status": "non-approved",
        }

    return output

def date_dir(myDir):
    yearDir = os.path.join(myDir, date_get()[0])
    dirControl = os.path.isdir(yearDir)
    if not dirControl:
        os.mkdir(yearDir)

    # mouthDir = os.path.join(yearDir, date_get()[1])
    # dirControl = os.path.isdir(mouthDir)
    # if not dirControl:
    #     os.mkdir(mouthDir)

    return yearDir

def download_file(
        goog_file_name,
        goog_file_id,
        file_location,
        month
    ):
    file_name = month + " " + goog_file_name
    local_folder = date_dir(file_location)

    full_file = os.path.join(local_folder, file_name)
    dirControl = os.path.isfile(full_file)
    if not dirControl:
        return goog_download(goog_file_id, file_name, local_folder)
    else:
        return full_file

# def download_regal(goog_file_name, goog_file_id):
#     file_name = date_get()[0] + " " + goog_file_name
#     local_folder = REGAL_LOCATION
#
#     full_file = os.path.join(local_folder, file_name)
#     dirControl = os.path.isfile(full_file)
#     if not dirControl:
#         return goog_download(goog_file_id, file_name, local_folder)
#     else:
#         return full_file

# def sai_file():
#     full_file = download_file(
#                             "Deconectari.xlsx",
#                             "1jK-zL0Xt38NALUprM9SW3-gEFba6kjTZ"
#                 )
#     wbDec = load_workbook(full_file)
#     wsNepr = wbDec["neprogramate-JT"]
#     wsProg = wbDec["programate-MT+JT"]
#
#     return (
#         wbDec,
#         wsNepr,
#         wsProg,
#         full_file
#     )

# def upload_dir():
#     myDir = LOCAL_DIRECTORY
#     dirControl = os.path.isdir(myDir)
#     if not dirControl:
#         os.mkdir(myDir)
#     myDir = myDir + "Uploads"
#     dirControl = os.path.isdir(myDir)
#     if not dirControl:
#         os.mkdir(myDir)
#
#     return myDir

# def upload_dir_empty():
#     myDir = UPLOAD_DIRECTORY
#     # for file in os.listdir(myDir):
#     #     os.remove(os.path.join(myDir, file))
#
#     return myDir

@app.route('/upload-file/', methods=['POST'])
def upload_al():
    file = request.files['file']
    filename = secure_filename(request.files['file'].filename)
    # full_file = "/".join([upload_dir_empty(), filename])
    # MYDIR =
    # goog_upload(file.read(), "filename")

    full_file = os.path.join(UPLOAD_DIRECTORY, filename)
    file.save(full_file)

    return full_file

@app.route('/upload-semn/', methods=['POST'])
def upload_semn():
    file = request.files['file']
    filename = secure_filename(request.files['file'].filename)
    full_file = os.path.join(UPLOAD_DIRECTORY, filename)
    if full_file.endswith(".png"):
        file.save(full_file)

        return full_file
    else:
        return "non-approved"

def goog_service():
    # -----------------------Lucrez cu google Drive API v3
    SCOPES = ['https://www.googleapis.com/auth/drive']
    creds = None
    # The file token.pickle stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists('google_my/token.pickle'):
        with open('google_my/token.pickle', 'rb') as token:
            creds = pickle.load(token)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'google_my/credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open('google_my/token.pickle', 'wb') as token:
            pickle.dump(creds, token)

    service = build('drive', 'v3', credentials=creds)

    return service

def goog_upload(full_file, file_name):

    # Partea care raspunde de upload
    file_metadata = {'name': file_name, 'parents': ['1TSSHbkN8EVfwKuyJ1RrSucmwrUSoEg46']}
    media = MediaFileUpload(full_file)
    file = goog_service().files().create(body=file_metadata,
                                  media_body=media,
                                  fields='id').execute()

    return file.get('id')

def goog_download(file_id, file_name, local_dir):
    request = goog_service().files().get_media(fileId=file_id)
    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while done is False:
        satus, done = downloader.next_chunk()

    fh.seek(0)
    full_file = os.path.join(local_dir, file_name)
    file = open(full_file, 'wb')
    file.write(fh.read())
    file.close()

    return full_file

@app.route('/inreg-al/<abrOf>/<dsAlNum>/<cu_dec>/<sef>/<mem>/<inst>/<inst_ad>/<pt>/<fid>/<lucr>/<sm>/<linkComma>/<autoSelect>/<tsSelect>/<dsDate>/', methods=['PUT'])
def get_inreg_al(abrOf, dsAlNum, cu_dec, sef, mem, inst, inst_ad, pt, fid, lucr, sm, linkComma, autoSelect, tsSelect, dsDate):
    if mem == "undefined" or mem == "Nu completati":
        mem = ""
    if inst == "undefined" or inst == "Nu completati":
        inst = ""
    if inst_ad != "undefined":
        if inst != "":
            inst = inst + ",\n" + inst_ad
        else:
            inst = inst_ad
    if pt == "undefined" or pt == "Nu completati":
        pt = ""
    if fid == "undefined" or fid == "0,":
        fid = ""
    else:
        fid = "F" + fid.strip(",")
    if sm == "undefined" or cu_dec == "Fara deconectari" or cu_dec == "Dispozitie" or sm == "0":
        sm = ""
    else:
        if cu_dec != "Fara deconectare" or cu_dec != "Dispozitie":
            sm = pt + "\n" + "dec. " + fid + ",\n" + "mont. SM" + sm

    reg_al = "al_" + date_get()[0] + "_" + date_get()[1]
    for i in mongo.db[reg_al].find().sort("_id", -1).limit(1):
        nrDsAl = i["_id"]

    try:
        nrDsAl = nrDsAl + 1
    except UnboundLocalError:
        try:
            prevMonth = datetime.datetime.now() - datetime.timedelta(days=28)
            prevMonth = prevMonth.strftime('%m')
            reg_al_prev = "al_" + date_get()[0] + "_" + prevMonth

            for i in mongo.db[reg_al_prev].find().sort("_id", -1).limit(1):
                nrDsAl = i["_id"]
            nrDsAl = nrDsAl + 1
        except UnboundLocalError:
            nrDsAl = 1

    if pt != "":
        pt_list = pt.split(',')
        pt_reg = "pt_" + abrOf.lower()
        for i in mgGen.db[pt_reg].find():
            if i["pt"] == pt_list[0]:
                localitatea = i["localitatea"]
                break

    else:
        localitatea = ""

    em = request.get_json()["emitent"]
    starea = request.get_json()["starea"]

    # if pos == "Dispecer":
    #     starea = "Acceptat:" + "\n" + date_get()[2] + "." + date_get()[1] + \
    #              "." + date_get()[5] + "\n" + date_get()[3] + ":" + date_get()[4] +"\n" + angaj
    # else:
    #     em = "Confirmat:" + "\n" + date_get()[2] + "." + date_get()[1] + \
    #              "." + date_get()[5] + "\n" + date_get()[3] + ":" + date_get()[4] +"\n" + em
    #
    #     starea = "Nou inregistrata"

    # Lucrez cu linkul de pe google
    if linkComma != "undefined" and autoSelect == "undefined":
        if re.search(",", linkComma):
            linkSlash = linkComma.replace(",", "/")
            linkSlash_list = linkSlash.split("/")
            for i in linkSlash_list:
                fileName = i
            googleLink = "https://drive.google.com/file/d/" + goog_upload(linkSlash, fileName)
        else:
            googleLink = "https://drive.google.com/file/d/" + linkComma

    elif autoSelect == "true":
        if tsSelect == "undefined":
            tsSelect = "Fara deconectari,De respectat distanta minim admisa (Tabelul nr. 1 NEI),De respectat cerintele CNESP COVID 19,De folosit mijloace individuale de protectie"

        if dsDate == "undefined":
            dsDate = date_get()[2] + "." + date_get()[1] + "." + date_get()[0]

        if dsAlNum == "undefined":
            fileName = str(nrDsAl) + ".pdf"
        else:
            fileName = str(dsAlNum) + ".pdf"
        linkSlash = os.path.join(UPLOAD_DIRECTORY, fileName)

        reg_angaj = "angajati_" + abrOf.lower()
        sef_list = sef.split()
        sef_name = sef_list[0] + " " + sef_list[1]
        emControl = re.search("Confirmat:", em)
        if emControl:
            em_list_conf = em.split("\n")
            # print(em_list_conf)
            em_list = em_list_conf[3].split()
        else:
            em_list = em.split()
        # print(em_list)
        em_name = em_list[0] + " " + em_list[1]
        em_name_gr = em_name + " gr. " + em_list[3]

        pdf = FPDF()
        pdf.add_page()

        pdf.set_font('Arial', 'B', 12)
        # pdf.set_text_color(2, 110, 212)
        pdf.cell(10, 10, "S.A. RED-Nord")

        pdf.ln(4)
        pdf.set_font('Arial', 'I', 10)
        pdf.set_text_color(0, 0, 0)
        pdf.cell(37, 10, f'ORED {mgGen.db.oficii.find_one({"abr": abrOf})["name"]}, sector ')
        pdf.set_font('Arial', 'IU', 10)
        pdf.set_text_color(200, 0, 0)
        pdf.cell(10, 10, mgGen.db[reg_angaj].find_one({"name": sef_name})["sector"])

        pdf.ln(20)
        pdf.set_font('Arial', 'B', 18)
        pdf.set_text_color(52, 107, 194)
        pdf.cell(42)
        pdf.cell(89, 10, "DISPOZIATIA DE LUCRU NR. ")
        pdf.set_font('Arial', 'IUB', 18)
        pdf.set_text_color(200, 0, 0)
        if dsAlNum == "undefined":
            pdf.cell(10, 10, str(nrDsAl))
        else:
            pdf.cell(10, 10, str(dsAlNum))

        pdf.ln(20)
        pdf.set_font('Arial', 'B', 12)
        pdf.set_text_color(0, 0, 0)
        pdf.cell(45, 10, "Sef de lucrari: ")
        pdf.ln(4)
        pdf.set_font('Arial', 'IU', 10)
        pdf.set_text_color(200, 0, 0)
        pdf.cell(30)
        pdf.cell(10, 10, sef)
        pdf.line(10, 66, 200, 66)

        pdf.ln(10)
        pdf.set_font('Arial', 'B', 12)
        pdf.set_text_color(0, 0, 0)
        pdf.cell(45, 10, "Membrii formatiei: ")

        pdf.ln(4)
        if mem != "":
            if mem[len(mem) - 1] != ",":
                mem += ","
        mem_list = mem.split(",")
        pdf.set_font('Arial', 'IU', 10)
        pdf.set_text_color(200, 0, 0)
        for i in mem_list:
            pdf.cell(30)
            pdf.cell(10, 10, i)
            pdf.ln(4)
        pdf.line(10, 76 + 4 * (len(mem_list) - 1), 200, 76 + 4 * (len(mem_list) - 1))

        pdf.ln(3)
        pdf.set_font('Arial', 'B', 12)
        pdf.set_text_color(0, 0, 0)
        pdf.cell(45, 10, "Sa se execute: ")

        pdf.ln(4)
        pdf.set_font('Arial', '', 10)
        pdf.set_text_color(0, 0, 0)
        pdf.cell(25)
        pdf.cell(40, 10, "Instalatia:")
        pdf.cell(25, 10, "PT:")
        pdf.cell(35, 10, "Fider nr.:")
        pdf.cell(35, 10, "Lucrari:")

        pdf.ln(4)
        pdf.cell(35)
        pdf.set_font('Arial', 'IU', 10)
        pdf.set_text_color(200, 0, 0)
        inst_list = inst.split(',')
        for i in inst_list:
            pdf.cell(10, 10, i.strip())
            pdf.ln(4)
            pdf.cell(35)

        setY = 87 + 4 * (len(mem_list) - 1)
        pdf.set_y(setY)
        if pt != "":
            if pt[len(pt) - 1] != ",":
                pt += ","
        pt_list = pt.split(',')
        for i in pt_list:
            pdf.cell(70)
            pdf.cell(10, 10, i)
            pdf.ln(4)
            # pdf.cell(10)

        # setY = 89 + 4 * (len(mem_list) - 1)
        pdf.set_y(setY)
        pdf.cell(100)
        pdf.cell(10, 10, fid.strip(','))

        if lucr[len(lucr) - 1] != ",":
            lucr += ","
        lucr_list = lucr.split(',')
        pdf.set_y(setY)
        for i in lucr_list:
            pdf.cell(133)
            pdf.cell(10, 10, i)
            pdf.ln(4)
            # pdf.cell(2)

        max_exec = max([len(inst_list), len(pt_list), len(lucr_list)])
        if inst_ad == "undefined" or max_exec != len(inst_list):
            line_exec_y = 4 * (max_exec - 1)
        else:
            line_exec_y = 4 * (max_exec - 1) + 4
        exec_y = 4 * (len(mem_list) - 1) + line_exec_y
        pdf.line(10, 91 + exec_y, 200, 91 + exec_y)

        # pdf.ln(10)
        setY = 94 + exec_y
        pdf.set_y(setY)
        pdf.set_font('Arial', 'B', 12)
        pdf.set_text_color(0, 0, 0)
        pdf.cell(45, 10, "Cerinte ale tehnicii securitatii: ")
        pdf.ln(4)
        tsSelect_list = tsSelect.split(",")
        for i in tsSelect_list:
            pdf.set_font('Arial', 'IU', 10)
            pdf.set_text_color(200, 0, 0)
            pdf.cell(30)
            pdf.cell(10, 10, i)
            pdf.ln(4)

        ts_y = exec_y + 4 * (len(tsSelect_list))
        pdf.line(10, 102 + ts_y, 200, 102 + ts_y)

        pdf.ln(7)
        pdf.set_font('Arial', 'B', 12)
        pdf.set_text_color(0, 0, 0)
        pdf.cell(72, 10, "Dispozitia a fost emisa pentru data de: ")
        pdf.ln(4)
        pdf.set_font('Arial', 'IU', 10)
        pdf.set_text_color(200, 0, 0)
        pdf.cell(30)
        pdf.cell(10, 10, dsDate)
        pdf.line(10, 117 + ts_y, 200, 117 + ts_y)

        pdf.ln(11)
        pdf.set_font('Arial', 'B', 12)
        pdf.set_text_color(0, 0, 0)
        pdf.cell(45, 10, "Emitentul dispozitiei: ")
        pdf.ln(4)
        pdf.cell(30)
        pdf.set_font('Arial', '', 10)
        pdf.cell(60, 10, "Nume, Prenume: ")
        pdf.cell(10, 10, "Semnatura: ")
        pdf.ln(4)
        pdf.set_font('Arial', 'IU', 10)
        pdf.set_text_color(200, 0, 0)
        pdf.cell(42)
        pdf.cell(10, 10, em_name_gr)

        semnatura_el = mgGen.db[reg_angaj].find_one({"name": em_name})["semnatura_el"]
        if semnatura_el != "":
            file_name = em_name + "_semn.png"

            setX = 120
            setY = 119 + ts_y
            pdf.image(goog_download(semnatura_el, file_name, UPLOAD_DIRECTORY), x=setX, y=setY, w=30, h=20)

        pdf.line(10, 136 + ts_y, 200, 136 + ts_y)

        pdf.ln(24)
        pdf.set_text_color(0, 0, 0)
        pdf.set_font('Arial', '', 10)
        pdf.cell(69, 10, "Permisiunea la pregatirea locului de munca")
        pdf.set_font('Arial', 'I', 10)
        pdf.cell(50, 10, "(data, ora):")
        pdf.line(98, 159 + ts_y, 128, 159 + ts_y)
        pdf.set_font('Arial', '', 10)
        pdf.cell(15, 10, "Dispecer")
        pdf.set_font('Arial', 'I', 10)
        pdf.cell(10, 10, "(nume):")
        pdf.line(158, 159 + ts_y, 200, 159 + ts_y)

        pdf.ln(4)
        pdf.ln(4)
        pdf.set_text_color(0, 0, 0)
        pdf.set_font('Arial', '', 10)
        pdf.cell(70, 10, "Permisiunea la admiterea membrilor echipei")
        pdf.set_font('Arial', 'I', 10)
        pdf.line(99, 167 + ts_y, 129, 167 + ts_y)
        pdf.cell(51, 10, "(data, ora):")
        pdf.set_font('Arial', '', 10)
        pdf.cell(15, 10, "Dispecer")
        pdf.set_font('Arial', 'I', 10)
        pdf.cell(10, 10, "(nume):")
        pdf.line(160, 167 + ts_y, 200, 167 + ts_y)

        pdf.ln(4)
        pdf.ln(4)
        pdf.set_text_color(0, 0, 0)
        pdf.set_font('Arial', '', 10)
        pdf.cell(69, 10, "Membrii formatiei au fost supusi instructajului la admitere la locul de lucru pregatit de catre:")

        pdf.ln(4)
        pdf.ln(4)
        pdf.set_text_color(0, 0, 0)
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(23, 10, "Sef de lucrari")
        pdf.set_font('Arial', 'I', 10)
        pdf.cell(50, 10, "(semnatura):")
        pdf.line(55, 183 + ts_y, 82, 183 + ts_y)

        pdf.set_text_color(0, 0, 0)
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(27, 10, "Membrii echipa")
        pdf.set_font('Arial', 'I', 10)
        pdf.cell(50, 10, "(semnatura):")
        pdf.line(133, 183 + ts_y, 200, 183 + ts_y)

        pdf.ln(4)
        pdf.ln(4)
        pdf.set_text_color(0, 0, 0)
        pdf.set_font('Arial', '', 10)
        pdf.cell(69, 10, "Lucrarile s-au terminat definitiv, formatia de lucru a fost evacuata:")

        pdf.ln(4)
        pdf.ln(4)
        pdf.set_text_color(0, 0, 0)
        pdf.set_font('Arial', '', 10)
        pdf.cell(35, 10, "Ora:")
        pdf.line(19, 199 + ts_y, 43, 199 + ts_y)
        # pdf.set_font('Arial', 'I', 10)
        # pdf.cell(50, 10, "(data, ora):")
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(23, 10, "Sef de lucrari")
        pdf.set_font('Arial', 'I', 10)
        pdf.cell(10, 10, "(semnatura):")
        pdf.line(91, 199 + ts_y, 140, 199 + ts_y)

        pdf.output(linkSlash)

        googleLink = "https://drive.google.com/file/d/" + goog_upload(linkSlash, fileName)
    else:
        googleLink = ""

    pt = pt.replace(",", ",\n")


    if dsAlNum == "undefined":
        mongo.db[reg_al].insert_one({
            "_id": nrDsAl,
            "oficiul": abrOf,
            "nr_ds": "",
            "nr_al": nrDsAl,
            "cu_dec": cu_dec,
            "instalatia": inst,
            "pt": pt.strip(",\n"),
            "fid_nr": fid,
            "localitatea": localitatea,
            "lucrarile": lucr.strip(","),
            "sef": sef,
            "mem_ech": mem.strip(","),
            "emitent": em,
            "mas_teh": sm,
            "semnatura": "Semnatura",
            "starea": starea,
            "pregatire": "Pregatire",
            "admitere": "Admitere",
            "terminare": "Terminare",
            "link": googleLink,
        })
    else:
        mongo.db[reg_al].update_one({"_id": int(dsAlNum)}, {"$set": {
            "oficiul": abrOf,
            "cu_dec": cu_dec,
            "instalatia": inst,
            "pt": pt.strip(",\n"),
            "fid_nr": fid,
            "localitatea": localitatea,
            "lucrarile": lucr.strip(","),
            "sef": sef,
            "mem_ech": mem.strip(","),
            "emitent": em,
            "mas_teh": sm,
            "semnatura": "Semnatura",
            "starea": starea,
            "pregatire": "Pregatire",
            "admitere": "Admitere",
            "terminare": "Terminare",
            "link": googleLink,
        }
        })
    return "true"

# @app.route('/inreg-ds/<abrOf>/<dsAlNum>/<lucr>/<sef>/<mem>/<inst>/<inst_ad>/<pt>/<fid>/<linkComma>/<autoSelect>/<tsSelect>/', methods=['PUT'])
# def get_inreg_ds(abrOf, dsAlNum, lucr, sef, mem, inst, inst_ad, pt, fid, linkComma, autoSelect, tsSelect):
#     if mem == "undefined":
#         mem = ""
#     if inst == "undefined":
#         inst = ""
#     if inst_ad != "undefined":
#         if inst != "":
#             inst = inst + ",\n" + inst_ad
#         else:
#             inst = inst_ad
#     if pt == "undefined":
#         pt = ""
#     if fid == "undefined":
#         fid = ""
#     else:
#         fid = "F" + fid.strip(",")
#     if tsSelect == "undefined":
#         tsSelect = "Fara deconectari,De respectat distanta minim admisa (Tabelul nr. 1 NEI),De respectat cerintele CNESP COVID 19,De folosit mijloace individuale de protectie"
#     # if linkComma == "undefined":
#     #     linkComma = ""
#
#     reg_al = "al_" + date_get()[0] + "_" + date_get()[1]
#     for i in mongo.db[reg_al].find().sort("_id", -1).limit(1):
#         nrDsAl = i["_id"]
#
#     try:
#         nrDsAl = nrDsAl + 1
#     except UnboundLocalError:
#         try:
#             prevMonth = datetime.datetime.now() - datetime.timedelta(days=28)
#             prevMonth = prevMonth.strftime('%m')
#             reg_al_prev = "al_" + date_get()[0] + "_" + prevMonth
#
#             for i in mongo.db[reg_al_prev].find().sort("_id", -1).limit(1):
#                 nrDsAl = i["_id"]
#             nrDsAl = nrDsAl + 1
#         except UnboundLocalError:
#             nrDsAl = 1
#
#     if pt != "":
#         pt_list = pt.split(',')
#         pt_reg = "pt_" + abrOf.lower()
#         for i in mgGen.db[pt_reg].find():
#             if i["pt"] == pt_list[0]:
#                 localitatea = i["localitatea"]
#                 break
#         pt = pt.replace(",", ",\n")
#     else:
#         localitatea = ""
#
#     em = request.get_json()["emitent"]
#     starea = request.get_json()["starea"]
#
#     if linkComma != "undefined":
#         linkSlash = linkComma.replace(",", "/")
#         linkSlash_list = linkSlash.split("/")
#         for i in linkSlash_list:
#             fileName = i
#         googleLink = "https://drive.google.com/file/d/" + goog_upload(linkSlash, fileName) + "/view?usp=sharing"
#     elif autoSelect == "true":
#         fileName = str(nrDsAl) + ".pdf"
#         linkSlash = os.path.join(UPLOAD_DIRECTORY, fileName)
#
#         reg_angaj = "angajati_" + abrOf.lower()
#         sef_list = sef.split()
#         sef_name = sef_list[0] + " " + sef_list[1]
#         emControl = re.search("Confirmat:", em)
#         if emControl:
#             em_list_conf = em.split("\n")
#             em_list = em_list_conf[3].split()
#         else:
#             em_list = em.split()
#         em_name = em_list[0] + " " + em_list[1]
#         em_name_gr = em_name + " gr. " + em_list[3]
#
#         pdf = FPDF()
#         pdf.add_page()
#
#         pdf.set_font('Arial', 'B', 12)
#         # pdf.set_text_color(2, 110, 212)
#         pdf.cell(10, 10, "S.A. RED-Nord")
#
#         pdf.ln(4)
#         pdf.set_font('Arial', 'I', 10)
#         pdf.set_text_color(0, 0, 0)
#         pdf.cell(37, 10, f'ORED {mgGen.db.oficii.find_one({"abr": abrOf})["name"]}, sector ')
#         pdf.set_font('Arial', 'IU', 10)
#         pdf.set_text_color(200, 0, 0)
#         pdf.cell(10, 10, mgGen.db[reg_angaj].find_one({"name": sef_name})["sector"])
#
#         pdf.ln(40)
#         pdf.set_font('Arial', 'B', 18)
#         pdf.set_text_color(52, 107, 194)
#         pdf.cell(42)
#         pdf.cell(89, 10, "DISPOZIATIA DE LUCRU NR. ")
#         pdf.set_font('Arial', 'IUB', 18)
#         pdf.set_text_color(200, 0, 0)
#         pdf.cell(10, 10, str(nrDsAl))
#
#         pdf.ln(25)
#         pdf.set_font('Arial', 'B', 12)
#         pdf.set_text_color(0, 0, 0)
#         pdf.cell(45, 10, "Sef de lucrari: ")
#         pdf.ln(4)
#         pdf.set_font('Arial', 'IU', 10)
#         pdf.set_text_color(200, 0, 0)
#         pdf.cell(30)
#         pdf.cell(10, 10, sef)
#         pdf.line(10, 92, 200, 92)
#
#         pdf.ln(13)
#         pdf.set_font('Arial', 'B', 12)
#         pdf.set_text_color(0, 0, 0)
#         pdf.cell(45, 10, "Membrii formatiei: ")
#
#         pdf.ln(4)
#         mem_list = mem.split(",")
#         pdf.set_font('Arial', 'IU', 10)
#         pdf.set_text_color(200, 0, 0)
#         for i in mem_list:
#             pdf.cell(30)
#             pdf.cell(10, 10, i)
#             pdf.ln(4)
#         pdf.line(10, 105 + 4 * (len(mem_list) - 1), 200, 105 + 4 * (len(mem_list) - 1))
#
#         pdf.ln(5)
#         pdf.set_font('Arial', 'B', 12)
#         pdf.set_text_color(0, 0, 0)
#         pdf.cell(45, 10, "Sa se execute: ")
#
#         pdf.ln(4)
#         pdf.set_font('Arial', '', 10)
#         pdf.set_text_color(0, 0, 0)
#         pdf.cell(25)
#         pdf.cell(40, 10, "Instalatia:")
#         pdf.cell(25, 10, "PT:")
#         pdf.cell(35, 10, "Fider nr.:")
#         pdf.cell(35, 10, "Lucrari:")
#
#
#         pdf.ln(4)
#         pdf.cell(35)
#         pdf.set_font('Arial', 'IU', 10)
#         pdf.set_text_color(200, 0, 0)
#         inst_list = inst.split(',')
#         for i in inst_list:
#             pdf.cell(10, 10, i.strip())
#             pdf.ln(4)
#             pdf.cell(35)
#
#         setY = 117 + 4 * (len(mem_list) - 1)
#         pdf.set_y(setY)
#         pt_list = pt.split(',')
#         for i in pt_list:
#             pdf.cell(70)
#             pdf.cell(10, 10, i)
#             pdf.ln(4)
#             # pdf.cell(10)
#
#         setY = 117 + 4 *(len(mem_list) - 1)
#         pdf.set_y(setY)
#         pdf.cell(100)
#         pdf.cell(10, 10, fid.strip(','))
#
#         lucr_list = lucr.split(',')
#         pdf.set_y(setY)
#         for i in lucr_list:
#             pdf.cell(133)
#             pdf.cell(10, 10, i)
#             pdf.ln(4)
#             # pdf.cell(2)
#
#         max_exec = max([len(inst_list), len(pt_list), len(lucr_list)])
#         if inst_ad == "undefined" or max_exec != len(inst_list):
#             line_exec_y = 4 * (max_exec - 1)
#         else:
#             line_exec_y = 4 * (max_exec - 1) + 4
#         exec_y = 4 * (len(mem_list) - 1) + line_exec_y
#         pdf.line(10, 122 + exec_y, 200, 122 + exec_y)
#
#         pdf.ln(10)
#         setY = 127 + exec_y
#         pdf.set_y(setY)
#         pdf.set_font('Arial', 'B', 12)
#         pdf.set_text_color(0, 0, 0)
#         pdf.cell(45, 10, "Cerinte ale tehnicii securitatii: ")
#         pdf.ln(4)
#         tsSelect_list = tsSelect.split(",")
#         for i in tsSelect_list:
#             pdf.set_font('Arial', 'IU', 10)
#             pdf.set_text_color(200, 0, 0)
#             pdf.cell(30)
#             pdf.cell(10, 10, i)
#             pdf.ln(4)
#
#         ts_y = exec_y + 4 * (len(tsSelect_list))
#         pdf.line(10, 136 + ts_y, 200, 136 + ts_y)
#
#         pdf.ln(10)
#         pdf.set_font('Arial', 'B', 12)
#         pdf.set_text_color(0, 0, 0)
#         pdf.cell(72, 10, "Dispozitia a fost eminsa la data de: ")
#         pdf.ln(4)
#         pdf.set_font('Arial', 'IU', 10)
#         pdf.set_text_color(200, 0, 0)
#         pdf.cell(30)
#         pdf.cell(10, 10, date_get()[2] + "." + date_get()[1] + "." + date_get()[0])
#         pdf.line(10, 154 + ts_y, 200, 154 + ts_y)
#
#         pdf.ln(14)
#         pdf.set_font('Arial', 'B', 12)
#         pdf.set_text_color(0, 0, 0)
#         pdf.cell(45, 10, "Emitentul dispozitiei: ")
#         pdf.ln(4)
#         pdf.cell(30)
#         pdf.set_font('Arial', '', 10)
#         pdf.cell(60, 10, "Nume, Prenume: ")
#         pdf.cell(10, 10, "Semnatura: ")
#         pdf.ln(4)
#         pdf.set_font('Arial', 'IU', 10)
#         pdf.set_text_color(200, 0, 0)
#         pdf.cell(42)
#         pdf.cell(10, 10, em_name_gr)
#
#         semnatura_el = mgGen.db[reg_angaj].find_one({"name": em_name})["semnatura_el"]
#         if semnatura_el != "":
#             file_name = em_name + "_semn.png"
#
#             setX = 120
#             setY = 158 + ts_y
#             pdf.image(goog_download(semnatura_el, file_name, UPLOAD_DIRECTORY), x=setX, y=setY, w=30, h=20)
#
#         pdf.line(10, 176 + ts_y, 200, 176 + ts_y)
#
#         pdf.output(linkSlash)
#
#         googleLink = "https://drive.google.com/file/d/" + goog_upload(linkSlash, fileName) + "/view?usp=sharing"
#     else:
#         googleLink = ""
#
#     if dsAlNum == "undefined":
#         mongo.db[reg_al].insert_one({
#             "_id": nrDsAl,
#             "oficiul": abrOf,
#             "nr_ds": nrDsAl,
#             "nr_al": "",
#             "cu_dec": "",
#             "instalatia": inst,
#             "pt": pt.strip(",\n"),
#             "fid_nr": fid,
#             "localitatea": localitatea,
#             "lucrarile": lucr.strip(","),
#             "sef": sef,
#             "mem_ech": mem.strip(","),
#             "emitent": em,
#             "mas_teh": "",
#             "semnatura": "Semnatura",
#             "starea": starea,
#             "pregatire": "Pregatire",
#             "admitere": "Admitere",
#             "terminare": "Terminare",
#             "link": googleLink,
#         })
#     else:
#         mongo.db[reg_al].update_one({"_id": int(dsAlNum)}, {"$set": {
#             "oficiul": abrOf,
#             "cu_dec": "",
#             "instalatia": inst,
#             "pt": pt.strip(",\n"),
#             "fid_nr": fid,
#             "localitatea": localitatea,
#             "lucrarile": lucr.strip(","),
#             "sef": sef,
#             "mem_ech": mem.strip(","),
#             "emitent": em,
#             "mas_teh": "",
#             "semnatura": "Semnatura",
#             "starea": starea,
#             "pregatire": "Pregatire",
#             "admitere": "Admitere",
#             "terminare": "Terminare",
#             "link": googleLink,
#         }
#         })
#
#     return "true"

@app.route('/inreg-sig/<abrOf>/<cauza_sig>/<cont>/<nume>/<adresa>/<disp>/<angaj>/', methods=['PUT'])
def get_sig(abrOf, cauza_sig, cont, nume, adresa, disp, angaj):
    now_time = request.get_json()["now_date"]
    # now_time = now_time.replace("\n", " ")
    disp = "Acceptat:\n" + now_time + "\n" + disp

    reg_sig = "sigilii_" + date_get()[0]

    for i in mongo.db[reg_sig].find().sort("_id", -1).limit(1):
        nr_sig = i["_id"]

    try:
        nr_sig += 1
    except UnboundLocalError:
        nr_sig = 1

    mongo.db[reg_sig].insert_one({
        "_id": nr_sig,
        "oficiul": abrOf,
        "cauza_sig": cauza_sig.strip(","),
        "cont": cont,
        "nume_cons": nume,
        "adresa": adresa,
        "disp": disp,
        "exec": angaj,
        "starea": "Nou inregistrat"
    })

    return "approved", 200

    # print(nr_sig)
    # print(abrOf)
    # print(cauza_sig)
    # print(cont)
    # print(nume)
    # print(adresa)
    # print(disp)
    # print(angaj)
    # print(now_time)

@app.route('/inreg-dec/<abrOf>/<cauza_dec>/<pt>/<fid>/<data>/<timp>/', methods=['PUT'])
def get_dec_nepr(abrOf, cauza_dec, pt, fid, data, timp):
#-----------------Dermin ora deconectarii si delta deconectarii-------------------
    dec_time_str = data + " " + timp
    now_time_str = request.get_json()["now_date"]
    now_time_str = now_time_str.replace("\n", " ")
    dec_time = datetime.datetime.strptime(dec_time_str, "%d.%m.%y %H:%M")
    now_time = datetime.datetime.strptime(now_time_str, "%d.%m.%y %H:%M")
    delta_time = now_time - dec_time
    myList = str(delta_time).split()
    try:
        if int(myList[0]):
            myDateSplit = myList[2].split(":")
            myDayPlus = 24 * int(myList[0]) + int(myDateSplit[0])
            myDateStr = str(myDayPlus) + ":" + myDateSplit[1] + ":" + myDateSplit[2]
            myDeltaHour = myDateStr
    except ValueError:
        myDeltaHour = str(delta_time)

    myDeltaHour_list = myDeltaHour.split(":")
    # print(myDeltaHour_list)

#--------------------------inainte de a face inscrieri in tabele controlez daca dela este pozitiva
#-------------------------pentru ca unii mai incurca introducerea datelor si delta iese negativa---------
    if int(myDeltaHour_list[0]) >= 0:
        fid = "F" + fid
        termText = "Incadrat"
        compens = 0
        myMinute = int(myDeltaHour_list[0]) * 60 + int(myDeltaHour_list[1])

        if myMinute > 8 * 60:
            difMinute = myMinute - 8 * 60
            termText = "Depasit cu: " + str(round(difMinute / 60)) + "H " + myDeltaHour_list[1] + "min"
            if difMinute / 60 <= 3:
                k = 1
            elif (difMinute / 60 > 3) and (difMinute / 60 <= 6):
                k = 4
            elif (difMinute / 60 > 6) and (difMinute / 60 <= 9):
                k = 7
            elif (difMinute / 60 > 9):
                k = 10
            compens = round(0.01 * 160 * 2.04 * k, 2)

        reg_pt = "pt_" + abrOf.lower()
        localitatea = ""
        fazaNrCas = 0
        fazaNrEc = 0

        for i in mgGen.db[reg_pt].find():
            if pt == i["pt"]:
                localitatea = i["localitatea"]
                totNrCas = i["total_cas"]
                if cauza_dec != "Ars FU 10kV,":
                    fazaNrCas = round(totNrCas / 9)
                    if fazaNrCas > 25:
                        fazaNrCas = random.randrange(20, 30)
                else:
                    fazaNrCas = round(totNrCas / 3)

                totNrEc = i["total_non_cas"]
                if totNrEc <= 2 and totNrCas == 0:
                    fazaNrEc = 1
                elif totNrEc > 12:
                    fazaNrEc = random.randrange(1, 3)
                else:
                    fazaNrEc = round(totNrEc / 9)
                break

        anlz_reg = f'anlzan{date_get()[5]}n'
        dec_reg = f'decnepr_{date_get()[0]}_{date_get()[1]}'

        for i in mongo.db[anlz_reg].find().sort("_id", -1).limit(1):
            nrAnlz = i["_id"]

        try:
            nrAnlz = nrAnlz + 1
        except UnboundLocalError:
            nrAnlz = 1

        for i in mongo.db[dec_reg].find().sort("_id", -1).limit(1):
            nrDec = i["_id"]

        try:
            nrDec = nrDec + 1
        except UnboundLocalError:
            nrDec = 1

        mongo.db[anlz_reg].insert_one({
            "_id": nrAnlz,
            "oficiul": abrOf,
            "pt_fider": pt + " " + fid,
            "localitate": localitatea,
            "nr_cons": fazaNrCas + fazaNrEc,
            "ore": myDeltaHour,
            "nr_dec": 1
        })

        mongo.db[dec_reg].insert_one({
            "_id": nrDec,
            "oficiul": abrOf,
            "pt": pt,
            "fid_04kv": fid,
            "data_dec": dec_time_str,
            "data_conect": now_time_str,
            "durata": myDeltaHour,
            "cons_cas": fazaNrCas,
            "cons_ec": fazaNrEc,
            "total": fazaNrCas + fazaNrEc,
            "localitate": localitatea,
            "cauza": cauza_dec,
            "termen": termText,
            "compens": compens,
        })
        

        # cur.execute("""SELECT * FROM pg_tables WHERE SCHEMANAME='public'""")
        # isTable = False
        # for table_tuple in cur.fetchall():
        #     if table_tuple[1] == f'anlzan{date_get()[5]}n':
        #         isTable = True
        #         break
        # if not isTable:
        #     cur.execute(f"""CREATE TABLE anlzan{date_get()[5]}n (
        #                             anlzan{date_get()[5]}n_id serial PRIMARY KEY,
        #                             oficiul VARCHAR(32),
        #                             pt_fider VARCHAR(128),
        #                             localitate VARCHAR(128),
        #                             nr_cons NUMERIC,
        #                             ore VARCHAR(32),
        #                             nr_dec NUMERIC,
        #                             nr_regl VARCHAR(128),
        #                             compens NUMERIC
        #                         )""")
        #     conn.commit()

        # cur.execute(f"""INSERT INTO anlzan{date_get()[5]}n (
        #                     oficiul,
        #                     pt_fider,
        #                     localitate,
        #                     nr_cons,
        #                     ore,
        #                     nr_dec
        #                 ) VALUES (
        #                     '{abrOf}',
        #                     '{pt + " " + fid}',
        #                     '{localitatea}',
        #                     '{fazaNrCas + fazaNrEc}',
        #                     '{myDeltaHour}',
        #                     '{1}'
        #                 )"""
        # )
        # conn.commit()

        # cur.execute("""SELECT * FROM pg_tables WHERE SCHEMANAME='public'""")
        # isTable = False
        # for table_tuple in cur.fetchall():
        #     if table_tuple[1] == f"decnepr_{date_get()[0]}_{date_get()[1]}":
        #         isTable = True
        #         break
        # if not isTable:
        #     cur.execute(f"""CREATE TABLE decnepr_{date_get()[0]}_{date_get()[1]} (
        #                                                         decnepr_id serial PRIMARY KEY,
        #                                                         oficiul VARCHAR(32),
        #                                                         nr_ordine INT4,
        #                                                         pt VARCHAR(32),
        #                                                         fid_04kv VARCHAR(32),
        #                                                         data_dec VARCHAR(32),
        #                                                         data_conect VARCHAR(32),
        #                                                         durata VARCHAR(32),
        #                                                         cons_cas INT4,
        #                                                         cons_ec INT4,
        #                                                         total INT4,
        #                                                         localitate VARCHAR(32),
        #                                                         cauza VARCHAR(256),
        #                                                         termen VARCHAR(256),
        #                                                         compens NUMERIC            
        #                             )""")
        #     conn.commit()

        # cur.execute(f"""SELECT decnepr_id FROM decnepr_{date_get()[0]}_{date_get()[1]}
        #                                 ORDER BY decnepr_id DESC LIMIT 1""")
        # decnepr_id_tuple = cur.fetchall()
        # if len(decnepr_id_tuple) == 0:
        #     nrDec = 1
        # else:
        #     nrDec = decnepr_id_tuple[0][0] + 1

        # cur.execute(f"""INSERT INTO decnepr_{date_get()[0]}_{date_get()[1]} (
        #                                     oficiul, 
        #                                     nr_ordine, 
        #                                     pt, 
        #                                     fid_04kv, 
        #                                     data_dec, 
        #                                     data_conect, 
        #                                     durata, 
        #                                     cons_cas, 
        #                                     cons_ec, 
        #                                     total, 
        #                                     localitate, 
        #                                     cauza, 
        #                                     termen, 
        #                                     compens
        #                             ) VALUES (
        #                                     '{abrOf}',
        #                                     '{nrDec}',
        #                                     '{pt}',
        #                                     '{fid}',
        #                                     '{dec_time_str}',
        #                                     '{now_time_str}',
        #                                     '{myDeltaHour}',
        #                                     '{fazaNrCas}',
        #                                     '{fazaNrEc}',
        #                                     '{fazaNrCas + fazaNrEc}',
        #                                     '{localitatea}',
        #                                     '{cauza_dec}',
        #                                     '{termText}',
        #                                     '{compens}'
        #                             )"""
        #             )
        # conn.commit()

        ################# Lucrez cu excel ################

        # full_file = download_file(
        #     "Deconectari.xlsx",
        #     "1KlOQQNZROaqs8C4xAQX1-kUaYNMCg9nz",
        #     SAIDI_LOCATION
        # )
        # wbDec = load_workbook(full_file)
        # wsNepr = wbDec["neprogramate-JT"]
        #
        # max_row = wsNepr.max_row + 1

        # wsNepr.cell(row=max_row, column=1).value = \
        #     wsNepr.cell(row=max_row - 1, column=1).value + 1
        # wsNepr.cell(row=max_row, column=1).alignment = \
        #     Alignment(horizontal="center", vertical="center")
        #
        # wsNepr.cell(row=max_row, column=2).value = abrOf
        # wsNepr.cell(row=max_row, column=2).alignment = \
        #     Alignment(horizontal="center", vertical="center")
        #
        # wsNepr.cell(row=max_row, column=3).value = \
        #     date_get()[2] + "." + date_get()[1] + "." + date_get()[5]
        # wsNepr.cell(row=max_row, column=3).alignment = \
        #     Alignment(horizontal="center", vertical="center")
        #
        # wsNepr.cell(row=max_row, column=4).value = \
        #     pt
        # wsNepr.cell(row=max_row, column=4).alignment = \
        #     Alignment(horizontal="center", vertical="center")
        #
        # wsNepr.cell(row=max_row, column=5).value = \
        #     fid
        # wsNepr.cell(row=max_row, column=5).alignment = \
        #     Alignment(horizontal="center", vertical="center")
        #
        # wsNepr.cell(row=max_row, column=6).value = \
        #     0
        # wsNepr.cell(row=max_row, column=6).alignment = \
        #     Alignment(horizontal="center", vertical="center")
        #
        # wsNepr.cell(row=max_row, column=7).value = \
        #     0
        # wsNepr.cell(row=max_row, column=7).alignment = \
        #     Alignment(horizontal="center", vertical="center")
        #
        # wsNepr.cell(row=max_row, column=8).value = \
        #     1
        # wsNepr.cell(row=max_row, column=8).alignment = \
        #     Alignment(horizontal="center", vertical="center")
        #
        # wsNepr.cell(row=max_row, column=9).value = \
        #     1
        # wsNepr.cell(row=max_row, column=9).alignment = \
        #     Alignment(horizontal="center", vertical="center")
        #
        # wsNepr.cell(row=max_row, column=10).value = \
        #     dec_time_str
        # wsNepr.cell(row=max_row, column=10).alignment = \
        #     Alignment(horizontal="center", vertical="center")
        #
        # wsNepr.cell(row=max_row, column=11).value = \
        #     now_time_str
        # wsNepr.cell(row=max_row, column=11).alignment = \
        #     Alignment(horizontal="center", vertical="center")
        #
        # wsNepr.cell(row=max_row, column=19).value = \
        #     myDeltaHour
        # wsNepr.cell(row=max_row, column=19).alignment = \
        #     Alignment(horizontal="center", vertical="center")
        #
        # wsNepr.cell(row=max_row, column=25).value = \
        #     fazaNrCas
        # wsNepr.cell(row=max_row, column=25).alignment = \
        #     Alignment(horizontal="center", vertical="center")
        #
        # wsNepr.cell(row=max_row, column=26).value = \
        #     fazaNrEc
        # wsNepr.cell(row=max_row, column=26).alignment = \
        #     Alignment(horizontal="center", vertical="center")
        #
        # wsNepr.cell(row=max_row, column=27).value = \
        #     fazaNrCas + fazaNrEc
        # wsNepr.cell(row=max_row, column=27).alignment = \
        #     Alignment(horizontal="center", vertical="center")
        #
        # wsNepr.cell(row=max_row, column=28).value = \
        #     localitatea
        # wsNepr.cell(row=max_row, column=28).alignment = \
        #     Alignment(horizontal="center", vertical="center")
        #
        # wsNepr.cell(row=max_row, column=32).value = \
        #     cauza_dec
        # # wsNepr.cell(row=max_row, column=32).alignment = \
        # #     Alignment(horizontal="center", vertical="center")

        # try:
        #     wbDec.save(full_file)
        # except (PermissionError):
        #     return "non_excel"

        return "approved"
    else:
        return "non_positive"
    

@app.route('/adaug-pt/<abrOf>/<pt>/<orSat>/<localitatea>/<sect>/<balanta>/<consCas>/<consNon>/')
def adaug_pt(abrOf, pt, orSat, localitatea, sect, balanta, consCas, consNon):
    # print(consCas)
    reg_pt = "pt_"+ abrOf.lower()
    for i in mgGen.db[reg_pt].find():
        if i["pt"] == pt:
            output = "repeated"
            break
        else:
            output = "success"
    if output == "success":
        for i in mgGen.db[reg_pt].find().sort("_id", -1).limit(1):
            idPt = i["_id"]

        idPt = idPt + 1
        localitatea = orSat + localitatea

        mgGen.db[reg_pt].insert_one({
            "_id": idPt,
            "pt": pt,
            "localitatea": localitatea,
            "balanta": balanta,
            "total_cas": int(consCas),
            "total_non_cas": int(consNon),
            "sector": sect,
        })

    return output

@app.route('/adaug-angaj/<abrOf>/<sect>/<tabel>/<name>/<pos>/<ts>/<pers>/<serv>/<em>/<cond>/<adm>/<sef>/<sup>/<mem>/<linkComma>/')
def adaug_angaj(abrOf, sect, tabel, name, pos, ts, pers, serv, em, cond, adm, sef, sup, mem, linkComma):
    if pers == "undefined":
        pers = ""
    if serv == "undefined":
        serv = ""
    if em == "undefined":
        em = ""
    if cond == "undefined":
        cond = ""
    if adm == "undefined":
        adm = ""
    if sef == "undefined":
        sef = ""
    if sup == "undefined":
        sup = ""
    if mem == "undefined":
        mem = ""

    if em == "false":
        em = "NU"
    elif em == "true":
        em = "DA"
    if cond == "false":
        cond = "NU"
    elif cond == "true":
        cond = "DA"
    if adm == "false":
        adm = "NU"
    elif adm == "true":
        adm = "DA"
    if sef == "false":
        sef = "NU"
    elif sef == "true":
        sef = "DA"
    if sup == "false":
        sup = "NU"
    elif sup == "true":
        sup = "DA"
    if mem == "false":
        mem = "NU"
    elif mem == "true":
        mem = "DA"

    if linkComma != "undefined":
        linkSlash = linkComma.replace(",", "/")
        linkSlash_list = linkSlash.split("/")
        for i in linkSlash_list:
            fileName = i
        fileId = goog_upload(linkSlash, fileName)
    else:
        fileId = ""



    reg_angaj = "angajati_" + abrOf.lower()

    for i in mgGen.db[reg_angaj].find().sort("_id", -1).limit(1):
        newId = i["_id"]

    newId = newId + 1
    mgGen.db[reg_angaj].insert_one({
        "_id": newId,
        "oficiul": abrOf,
        "sector": sect,
        "nr_tabel": tabel,
        "name": name,
        "position": pos,
        "gr_ts": ts,
        "telefon_pers": pers,
        "telefon_serv": serv,
        "emitent": em,
        "conducator": cond,
        "admitent": adm,
        "sef": sef,
        "supraveghetor": sup,
        "membru": mem,
        "semnatura_el": fileId,
    })

    return "true"

# @app.route('/inreg-deranj/<of_select>/<angaj_select>/<sect_select>/<inst-select>/<fid10_select>/<pt_select>/<fid04_select>/<cont_select>/')
@app.route('/inreg-deranj/<of_select>/<angaj_select>/<sect_select>/<inst_select>/<fid10_select>/<pt_select>/<fid04_select>/<cont_select>/')
def get_inreg_deranj(of_select, angaj_select, sect_select, inst_select, fid10_select, pt_select, fid04_select, cont_select):
    if angaj_select == "undefined":
        angaj_select = ""
    if fid10_select == "undefined":
        fid10_select = ""
    if pt_select == "undefined":
        pt_select = ""
    if fid04_select == "undefined":
        fid04_select = ""

    reg_deranj = "deranj_" + date_get()[0]
    for i in mongo.db[reg_deranj].find().sort("_id", -1).limit(1):
        nr_deranj = i["_id"]

    try:
        nr_deranj = nr_deranj + 1
    except UnboundLocalError:
        nr_deranj = 1

    mongo.db[reg_deranj].insert_one({
        "_id": nr_deranj,
        "oficiul": of_select,
        "nr_ordine": nr_deranj,
        "transmis": angaj_select,
        "sector": sect_select,
        "instalatia": inst_select,
        "fid_10kv": fid10_select,
        "pt": pt_select,
        "fid_04kv": "F" + fid04_select,
        "continutul": cont_select,
        "data": date_get()[6].strftime('%d.%m.%Y %H:%M'),
        "responsabil": "Semnatura",
        "starea": "Neexecutat"
    })

    return "true"

@app.route('/find-by-name/<abrOf>/<name>/')
def find_by_name(abrOf, name):
    output = []
    reg_angaj = "angajati_" + abrOf.lower()
    for i in mgGen.db[reg_angaj].find({"name": name}):
        output = {
            "sector": i["sector"],
            "nr_tabel": i["nr_tabel"],
            "position": i["position"],
            "gr_ts": i["gr_ts"],
            "telefon_pers": i["telefon_pers"],
            "telefon_serv": i["telefon_serv"],
            "emitent": i["emitent"],
            "conducator": i["conducator"],
            "admitent": i["admitent"],
            "sef": i["sef"],
            "supraveghetor": i["supraveghetor"],
            "membru": i["membru"],
            "semnatura_el": i["semnatura_el"],
        }

    return jsonify(output)

@app.route('/modif-angaj/<abrOf>/<name>/<tabel>/<sect>/<pos>/<ts>/<pers>/<serv>/<em>/<cond>/<adm>/<sef>/<sup>/<mem>/<linkComma>/')
def modif_angaj(abrOf, name, tabel, sect, pos, ts, pers, serv, em, cond, adm, sef, sup, mem, linkComma):
    if pers == "undefined":
        pers = ""
    if serv == "undefined":
        serv = ""
    if em == "undefined":
        em = ""
    if cond == "undefined":
        cond = ""
    if adm == "undefined":
        adm = ""
    if sef == "undefined":
        sef = ""
    if sup == "undefined":
        sup = ""
    if mem == "undefined":
        mem = ""

    if em == "false":
        em = "NU"
    elif em == "true":
        em = "DA"
    if cond == "false":
        cond = "NU"
    elif cond == "true":
        cond = "DA"
    if adm == "false":
        adm = "NU"
    elif adm == "true":
        adm = "DA"
    if sef == "false":
        sef = "NU"
    elif sef == "true":
        sef = "DA"
    if sup == "false":
        sup = "NU"
    elif sup == "true":
        sup = "DA"
    if mem == "false":
        mem = "NU"
    elif mem == "true":
        mem = "DA"

    if linkComma != "undefined":
        linkSlash = linkComma.replace(",", "/")
    else:
        linkSlash = ""

    reg_angaj = "angajati_" + abrOf.lower()

    mgGen.db[reg_angaj].update_one({"name": name}, {"$set": {
        "oficiul": abrOf,
        "sector": sect,
        "nr_tabel": tabel,
        "position": pos,
        "gr_ts": ts,
        "telefon_pers": pers,
        "telefon_serv": serv,
        "emitent": em,
        "conducator": cond,
        "admitent": adm,
        "sef": sef,
        "supraveghetor": sup,
        "membru": mem,
        "semnatura_el": linkSlash,
    }
    })

    return "true"

@app.route('/starea-sig/<sigId>/', methods=['PUT'])
def get_starea_sig(sigId):
    reg_sig = "sigilii_" + date_get()[0]

    if mongo.db[reg_sig].find_one({"_id": int(sigId)})["starea"] == "Nou inregistrat":
        mongo.db[reg_sig].update_one({"_id": int(sigId)}, {"$set": {"starea": request.get_json()["starea"]}})
        return "approved"
    else:
        return "non-approved", 406


@app.route('/lista-sig/<abrOf>/')
def get_lista_sig(abrOf):
    output = [{
        "_id": "",
        "oficiul": "",
        "cauza_sig": "",
        "cont": "",
        "nume_cons": "",
        "adresa": "",
        "disp": "",
        "exec": "",
        "starea": "",
    }]

    reg_sig = "sigilii_" + date_get()[0]
    if abrOf == "AD":
        for i in mongo.db[reg_sig].find():
            output.append(i)
    else:
        for i in mongo.db[reg_sig].find({"oficiul": abrOf}):
            output.append(i)

    return jsonify(output)

@app.route('/lista-pt/<abrOf>/<pt>/')
def get_lista_pt(abrOf, pt):
    # output = [{
    #     "pt": "",
    #     "localitatea": "",
    #     "balanta": "",
    #     "total_cas": "",
    #     "total_non_cas": "",
    #     "sector": ""
    # }]
    output = []
    if abrOf != "AD":
        listaPt = "pt_" + abrOf.lower()
        if pt == "undefined":
            for i in mgGen.db[listaPt].find():
                i["oficiul"] = abrOf
                output.append(i)
        elif pt != "undefined":
            for i in mgGen.db[listaPt].find({"pt": pt}):
                i["oficiul"] = abrOf
                output.append(i)
    else:
        for i in OFICII:
            listaPt = "pt_" + i["abr"].lower()
            if pt == "undefined":
                for j in mgGen.db[listaPt].find():
                    j["oficiul"] = i["abr"]
                    output.append(j)
            elif pt != "undefined":
                for j in mgGen.db[listaPt].find({"pt": pt}):
                    j["oficiul"] = i["abr"]
                    output.append(j)

            # print(output)

    output = sorted(output, key=lambda k: k['pt'])
    # pprint.pprint(output)
    return jsonify(output)

@app.route('/lista-angaj/<abrOf>/<sect>/')
def get_lista_angaj(abrOf, sect):
    output = []
    listaAngaj = "angajati_" + abrOf.lower()
    if sect == "undefined":
        for i in mgGen.db[listaAngaj].find():
            output.append(i)
    elif sect != "undefined":
        for i in mgGen.db[listaAngaj].find({"sector": sect}):
            output.append(i)

    output = sorted(output, key=lambda k: k['name'])
    return jsonify(output)

@app.route('/nr-tabel/<abrOf>/<angaj>/')
def get_tabel(abrOf, angaj):
    output = []
    lista_angaj = "angajati_" + abrOf.lower()

    for i in mgGen.db[lista_angaj].find({"name": angaj}):
        output.append(i)

    return jsonify(output)

@app.route('/sterg-angaj/<abrOf>/<angaj>/<nr_tabel>/')
def get_sterg_angaj(abrOf, angaj, nr_tabel):
    lista_angaj = "angajati_" + abrOf.lower()

    mgGen.db[lista_angaj].delete_one({"name": angaj, "nr_tabel": nr_tabel})

    return "true"

@app.route('/sterg-pt/<abrOf>/<pt>/')
def get_sterg_pt(abrOf, pt):
    lista_pt = "pt_" + abrOf.lower()

    mgGen.db[lista_pt].delete_one({"pt": pt})

    return "true"


# @app.route('/lista-angaj-ts/<abrOf>/<sect>/')
# def get_lista_angaj(abrOf, sect):
#     output = []
#     listaAngaj = "angajati_" + abrOf.lower()
#     if sect == "undefined":
#         for i in mongo.db[listaAngaj].find():
#             output.append(i)
#     elif sect != "undefined":
#         for i in mongo.db[listaAngaj].find({"sector": sect}):
#             output.append(i)
#
#     output = sorted(output, key=lambda k: k['name'])
#     return jsonify(output)






#
# @app.route('/generare-cod/<telefon_serv>/<telefon_pers>/')
# def auth_func(telefon_serv, telefon_pers):
#     if telefon_serv != "" or telefon_pers != "":
#         if re.match('[+]373 \d\d\d\d\d\d\d\d', telefon_serv):
#             twilioFunc(telefon_serv)
#             tel_verify = {"telefon": telefon_serv, "twilio": verify}
#             return jsonify(tel_verify)
#         elif re.match('[+]373 \d\d\d\d\d\d\d\d', telefon_pers):
#             twilioFunc(telefon_pers)
#             tel_verify = {"telefon": telefon_pers, "twilio": verify}
#             return jsonify(tel_verify)
#
# @app.route('/approved/<smsCode>/<myTel>/')
# def get_valid(smsCode, myTel):
#     # print(smsCode)
#     # print(myTel)
#     result = veryfy.verification_checks.create(to=myTel, code=smsCode)
#     if result.status == 'approved':
#         return "approved"






